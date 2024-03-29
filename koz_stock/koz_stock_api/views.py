from django.shortcuts import render
import pandas as pd
from .models import ( Products, ProductInflow, ProductInflowItem, ProductOutflow, Consumers, Suppliers,
                      Stock, Project, CustomUser, Company, ProductGroups,
                      SequenceNumber, ProductSubgroups, ProductInflowImage, ProductOutflowImage,
                      QuantityTakeOff, Building, Section, Place, ElevationOrFloor, AccountingInflow, AccountingItem)
from django.views import View
from rest_framework.views import APIView
from django.http import JsonResponse, HttpResponse
import json
from django.db.models.signals import post_save, post_delete, pre_save, pre_delete
from django.dispatch import receiver
from .definitions import create_receipt_pdf
from datetime import datetime
import datetime
from django.db.models import Sum, Avg, Max, F, ProtectedError
from django.db.models.functions import Coalesce
from django.views.decorators.csrf import csrf_exempt
import statistics
import math
from scipy.stats import norm
import traceback
from django.db import models
import openpyxl
import base64
from io import BytesIO
import numpy as np
from django.db import transaction, IntegrityError
from .permissions import IsSuperStaff, IsAccountingStaff, IsStockStaff, IsSuperStaffOrStockStaff, IsSuperStaffOrAccountingStaff
from django.utils.translation import gettext as _
import os
import tempfile
from django.http import FileResponse, Http404
import urllib.parse
from django.conf import settings
from django.conf.urls.static import static


# Authentications
from django.contrib.auth import authenticate
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.decorators import  permission_classes, authentication_classes
from django.db.utils import OperationalError
from rest_framework.exceptions import ValidationError
import filetype
from django.core.exceptions import ObjectDoesNotExist
from rest_framework.exceptions import NotAuthenticated, PermissionDenied
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.core.exceptions import ValidationError

from django.http import JsonResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from django.core.exceptions import ObjectDoesNotExist
from .models import Project
from django.utils import translation
from PIL import Image



# region Language

def set_language(request):
    user_language = 'tr'  # replace with the user's preferred language
    translation.activate(user_language)
    request.session[translation.LANGUAGE_SESSION_KEY] = user_language

# endregion


# region Login/Logout

class LoginView(TokenObtainPairView):
    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        username = data.get('username')
        username = username.strip()
        password = data.get('password')
       
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_active:
            response = super().post(request, *args, **kwargs)
            # Convert the first name and last name to uppercase
            first_name = user.first_name.upper()
            last_name = user.last_name.upper()
            # Add the uppercase names to the response
            response.data['first_name'] = first_name
            response.data['last_name'] = last_name
            return response
        else:
            return JsonResponse({'error': _('Invalid credentials')}, status=401)

class LogoutView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def post(self, request):
        try:
            refresh_token = request.POST.get('refresh_token')
            token = RefreshToken(refresh_token)
            token.blacklist()

            return JsonResponse({'success': _('successfully log out')}, status=205)
        except Exception as e:
            return JsonResponse({'error': _('BAD REQUEST')}, status=400)


# endregion

# region Project Selection

class CreateProjectView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        name = request.data.get('name')
        company = request.user.company  # Get the company from the logged in user

        if not name or not company:
            return JsonResponse({'error': _('Name and company are required')}, status=400)

        project = Project(name=name, company=company)
        project.save()

        for user in CustomUser.objects.filter(company=request.user.company):
            user.projects.add(project)
            user.save()
        
        for group in ProductGroups.objects.filter(company=request.user.company):
            group.projects.add(project)
            group.save()

        return JsonResponse({'message': _('Project created and assigned to all users successfully')})

class AddExcelProjectView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)
        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            if 'file' not in request.FILES:
                return JsonResponse({'error': _("No file uploaded")}, status=400)
            
            file = request.FILES['file']
            kind = filetype.guess(file.read())
            
            if kind is None or kind.mime not in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
                return JsonResponse({'error': _("The uploaded file is not a valid Excel file")}, status=400)

            data = pd.read_excel(file)
            if data.empty:
                return JsonResponse({'error': _("The uploaded file is empty")}, status=400)

            count = 0
            for i, row in data.iterrows():

                # Fetching or Creating based on hierarchy and names
                project, _ = Project.objects.get_or_create(name=row["Project Name"], company=request.user.company)

                building, _ = Building.objects.get_or_create(name=row["Building Name"], project=project) if "Building Name" in row and pd.notna(row["Building Name"]) else (None, None)

                elevation_or_floor, _ = ElevationOrFloor.objects.get_or_create(name=row["Elevation or Floor Name"], building=building) if "Elevation or Floor Name" in row and pd.notna(row["Elevation or Floor Name"]) else (None, None)

                section, _ = Section.objects.get_or_create(name=row["Section Name"], elevation_or_floor=elevation_or_floor) if "Section Name" in row and pd.notna(row["Section Name"]) else (None, None)

                place, _ = Place.objects.get_or_create(name=row["Place Name"], section=section) if "Place Name" in row and pd.notna(row["Place Name"]) else (None, None)

                count += 1

            return JsonResponse({'message': f"{count} Project hierarchy entries added successfully"}, status=200)

        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)

class GetProjectsView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def get(self, request):
        # Get the projects that belong to the user's company and the user is associated with
        projects = request.user.projects.filter(company=request.user.company)
        projects = list(projects.values('id', 'name'))  # Convert QuerySet to a list
        print(projects)
        return JsonResponse(projects, safe=False)

class GetCurrentProjectView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def get(self, request):
        # Get the projects that belong to the user's company and the user is associated with
        project = request.user.current_project
        project = [project.id ,project.name] # Convert QuerySet to a list
        
        return JsonResponse(project, safe=False)


class SetCurrentProjectView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request):
        data = json.loads(request.body)
        project_id = data.get('project_id')
        try:
            print("current_project:", request.user.current_project)
            print("project_id:", project_id)
            project = Project.objects.get(id=project_id, company=request.user.company)
            print(project)
            # Check if the project belongs to the user's projects
            if project not in request.user.projects.all():
                return JsonResponse({'error': _('You do not have access to this project.')}, status=403)

            request.user.current_project = project
            request.user.save()
            return JsonResponse({'error': f"{project.name}, {request.user.current_project.name},{project_id}"})
            #return JsonResponse({'message': _('Project set successfully')})
        except ObjectDoesNotExist:
            return JsonResponse({'error': _('Project not found or does not belong to your company')}, status=400, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


# endregion

# region User

class CreateUserView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            # Extract the data from the request
            print(request.data)
            username = request.data.get('username')
            password = request.data.get('password')
            first_name = request.data.get('first_name')
            last_name = request.data.get('last_name')
            staff_role = request.data.get('staff_role')
            project_ids = request.data.get('projects', [])
            # The new user's company should be equal to superstaff's company and set current_project to None
            company = request.user.company
            current_project = None

            # Check if the username is provided
            if not username:
                return JsonResponse({'error': _("Username is required.")}, status=400)

            # Check if the password is provided
            if not password:
                return JsonResponse({'error': _("Password is required.")}, status=400)

            # Check if the username is already taken
            if CustomUser.objects.filter(username=username, company=company).exists():
                return JsonResponse({'error': _("Username is already taken.")}, status=400)

            

            # Create the user
            user = CustomUser(
                username=username, 
                first_name=first_name, 
                last_name=last_name, 
            )
            if staff_role == "super_staff":
                user.is_superstaff = True
                user.is_stockstaff = False
                user.is_accountingstaff = False
            elif staff_role == "stock_staff":
                user.is_superstaff = False
                user.is_stockstaff = True
                user.is_accountingstaff = False
            elif staff_role == "accounting_staff":
                user.is_superstaff = False
                user.is_stockstaff = False
                user.is_accountingstaff = True
            user.set_password(password)
            user.company = company
            user.is_staff = True
            user.current_project = current_project
            user.save()

            # Assign the projects to the user
            projects = Project.objects.filter(id__in=project_ids, company=company)
            user.projects.set(projects)

            return JsonResponse({'message': _('User created successfully')})

        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class EditUserView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            # Extract the data from the request
            user_id = request.data.get('id')
            username = request.data.get('username')
            password = request.data.get('password')
            first_name = request.data.get('first_name')
            last_name = request.data.get('last_name')
            staff_role = request.data.get('staff_role')
            project_ids = request.data.get('projects', [])
            company = request.user.company

            # Check if the username is provided
            if not username:
                return JsonResponse({'error': _("Username is required.")}, status=400)

            # Check if the staff role is provided
            if not staff_role:
                return JsonResponse({'error': _("Staff role is required.")}, status=400)

            # Check if the staff role is valid
            if staff_role not in ["super_staff", "stock_staff", "accounting_staff"]:
                return JsonResponse({'error': _("Invalid staff role.")}, status=400)

            # Get the User, Projects instances
            user = CustomUser.objects.get(id=user_id, company=company)
            projects = Project.objects.filter(id__in=project_ids, company=company)

            # Update the user details
            if username:
                user.username = username
            if password:
                user.set_password(password)
            if first_name:
                user.first_name = first_name
            if last_name:
                user.last_name = last_name
            if staff_role == "super_staff":
                user.is_superstaff = True
                user.is_stockstaff = False
                user.is_accountingstaff = False
            elif staff_role == "stock_staff":
                user.is_superstaff = False
                user.is_stockstaff = True
                user.is_accountingstaff = False
            elif staff_role == "accounting_staff":
                user.is_superstaff = False
                user.is_stockstaff = False
                user.is_accountingstaff = True
            user.save()

            # Assign the projects to the user
            user.projects.set(projects)

            return JsonResponse({'message': 'User details updated successfully'})
        
        except CustomUser.DoesNotExist:
            return JsonResponse({'error': _("User not found.")}, status=400)

        except Project.DoesNotExist:
            return JsonResponse({'error': _("Project not found.")}, status=400)

        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)


class CollapsedUserView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)
    
    def get(self, request, *args, **kwargs):
        active_users = CustomUser.objects.filter(is_active=True)
        active_users_list = [[user.id, user.first_name, user.last_name, user.is_active] for user in active_users]
        
        return JsonResponse({"active_users_list": active_users_list,
                             }, safe=False)

class UserCardView(APIView):
    permission_classes = (IsAuthenticated,) 
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)
    
    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        id = data.get('id')
        print(id)
        try:
            user = CustomUser.objects.get(id=id)
        except ObjectDoesNotExist:
            return JsonResponse({'error': _('User not found')}, status=404)

        if user.is_superstaff:
            staff_role = "super_staff"
        elif user.is_stockstaff:
            staff_role = "stock_staff"
        elif user.is_accountingstaff:
            staff_role = "accounting_staff"
        else:
            staff_role = None
        # Check if the staff role is valid
        if staff_role not in ["super_staff", "stock_staff", "accounting_staff"]:
            return JsonResponse({'error': _("Invalid staff role.")}, status=400)
        response_data = {
            'id': user.id, 
            'username': user.username, 
            'first_name': user.first_name,
            'last_name': user.last_name,
            'is_superuser': user.is_superuser,
            'is_active': user.is_active,
            'staff_role': staff_role,
            'projects': [[project.id ,project.name] for project in user.projects.all()],
        }

        return JsonResponse(response_data, safe=False)

class DeleteUserView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        user_id = request.data.get('user_id')
        try:
            user = CustomUser.objects.get(id=user_id)
            if request.user.is_superstaff or request.user.is_superuser:
                user.delete()
                return JsonResponse({'message': _('User deleted successfully')})
            else:
                return JsonResponse({'error': _('You do not have permission to delete this user')}, status=403)
        except ProtectedError:
            return JsonResponse({
                'error': _("This User cannot be deleted because it is being used elsewhere in the application.")
            }, status=400)
        except CustomUser.DoesNotExist:
            return JsonResponse({'error': _('User not found')}, status=404)


# endregion


#region Products


class AddProductsView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def get(self, request, *args, **kwargs):
        try:
            groups = ProductGroups.objects.filter(company=request.user.company)
            subgroups = ProductSubgroups.objects.filter(group__company=request.user.company)

            group_names = list(groups.values_list('group_name', flat=True))
            subgroup_names = list(subgroups.values_list('subgroup_name', flat=True))

            return JsonResponse({
                'group_names': group_names,
                'subgroup_names': subgroup_names
            }, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            group_code = data.get('group_code')
            subgroup_code = str(data.get('subgroup_code'))
            print(group_code)
            print(subgroup_code)
            

            # Fetch group and subgroup objects using their names
            try:
                group = ProductGroups.objects.get(group_code=group_code, company=request.user.company)
                print(group)
                subgroup = ProductSubgroups.objects.get(subgroup_code=subgroup_code, group=group, group__company=request.user.company)
            except (ProductGroups.DoesNotExist, ProductSubgroups.DoesNotExist):
                error_message = _("The specified group or subgroup does not exist.")
                return JsonResponse({'error': error_message}, status=400)

            # Construct product code
            product_code = f'{group.group_code}.{subgroup.subgroup_code}.{subgroup.sequence_number}'

            # Increment sequence_number in the subgroup
            subgroup.sequence_number += 1
            subgroup.save()

            # Add new product
            new_product_data = {}
            for field in ['brand', 'serial_number', 'model', 'description', 'unit']:
                value = data.get(field)
                if value is not None and value != '':
                    new_product_data[field] = value
                else:
                    error_message = _("The field '{%s}' cannot be empty.") % field
                    return JsonResponse({'error': error_message}, status=400)

            new_product = Products.objects.create(product_code=product_code, group=group, subgroup=subgroup, company=request.user.company, **new_product_data)
            # Create a new Stock object for the current user's current project
            Stock.objects.create(
                product=new_product,
                warehouse=None,
                inflow=0,
                outflow=0,
                stock=0,
                company=request.user.company,
                project=request.user.current_project
            )

            message = _("New product '{%s}' has been successfully created.") % new_product.product_code
            return JsonResponse({'message': message}, status=201)

        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)



class ProductsView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def get(self, request, *args, **kwargs):
        company = request.user.company

        if not company:
            return JsonResponse({'error': _('Company not associated with user.')}, status=400)

        products = Products.objects.filter(company=company)
        product_list = [[p.id, p.product_code, [p.group.group_code, p.group.group_name], [p.subgroup.subgroup_code, p.subgroup.subgroup_name], p.brand,
                         p.serial_number, p.model, p.description, p.unit] for p in products]
        
        return JsonResponse(product_list, safe=False, status=200)


class EditProductsView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            id = data.get('id')
            product = Products.objects.get(id=id)

            # Check if product belongs to the same company as the user
            if product.company != request.user.company:
                return JsonResponse({'error': _("Product doesn't belong to your company!")}, status=400)

            # Update other product fields
            for field in  ['brand', 'serial_number', 'model', 'description', 'unit']:
                value = data.get(field)
                if value is not None and value != '':
                    setattr(product, field, value)

            new_group_id = data.get('group')
            if new_group_id:
                try:
                    group = ProductGroups.objects.get(group_code=new_group_id, company=request.user.company)
                    product.group = group
                except ProductGroups.DoesNotExist:
                    return JsonResponse({'error': _("Group with name '%s' does not exist.") % new_group_id}, status=400)

            new_subgroup_id = data.get('subgroup')
            if new_subgroup_id:
                try:
                    subgroup = ProductSubgroups.objects.get(subgroup_code=new_subgroup_id, group=group,  group__company=request.user.company)
                    product.subgroup = subgroup
                except ProductSubgroups.DoesNotExist:
                    traceback.print_exc()
                    return JsonResponse({'error': _("Subgroup with name '%s' does not exist.") % new_subgroup_id}, status=400)

            #product.refresh_from_db()
            
            dirty_fields = product.get_dirty_fields(check_relationship=True)

            if 'group' in dirty_fields or 'subgroup' in dirty_fields:
                # Here, assign the new product code based on the new group and subgroup
                new_product_code = f'{group.group_code}.{subgroup.subgroup_code}.{subgroup.sequence_number}'
                product.product_code = new_product_code
                subgroup.sequence_number += 1
                subgroup.save()
                
            product.save()
            return JsonResponse({'message': _("Your changes have been successfully saved.")}, status=200)

        except Products.DoesNotExist: 
            return JsonResponse({'error': _("Product not found.")}, status=400)

        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class DeleteProductsView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)
    
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            id = data.get('id')
            Products.objects.get(id=id).delete()
        except ProtectedError:
            return JsonResponse({
                'error': _("This product cannot be deleted because it is being used elsewhere in the application.")
            }, status=400)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)
        return JsonResponse({'message': _("Product object has been successfully deleted")}, status=200)



# endregion

# region suppliers

class AddSuppliersView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            user = request.user
            data = json.loads(request.body)
            tax_code = data.get('tax_code')
            if not tax_code:
                return JsonResponse({'error': _("Tax Code cannot be empty!")}, status=400)
            elif Suppliers.objects.filter(tax_code=tax_code, company=user.company).exists():
                error_message = _("The Tax Code '%s' already exists in the database for this company.") % tax_code
                return JsonResponse({'error': error_message}, status=400)

            # List of mandatory fields
            mandatory_fields = ['name', 'contact_name', 'contact_no']

            # Initialize a dictionary to store new supplier data
            new_supplier_data = {}

            for field in ['name', 'contact_name', 'contact_no', 'mail', 'address', 'explanation']:
                value = data.get(field)

                # If the value is not None and not an empty string, save it
                if value is not None and value != '':
                    new_supplier_data[field] = value

                # If the value is None or an empty string, and the field is mandatory, raise an error
                elif field in mandatory_fields:
                    error_message = _("The field '{%s}' cannot be empty.") % field
                    return JsonResponse({'error': error_message}, status=400)

            new_supplier = Suppliers.objects.create(tax_code=tax_code, company=user.company, **new_supplier_data)

            message = _("New supplier '{%s}' has been successfully created.") % new_supplier.name
            return JsonResponse({'message': message}, status=201)

        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class SuppliersView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def get(self, request, *args, **kwargs):
        company = request.user.company
        suppliers = Suppliers.objects.filter(company=company)
        supplier_list = [[s.id, s.tax_code, s.name, s.contact_name, s.contact_no, s.mail, s.adress, s.explanation] for s in suppliers]
        return JsonResponse(supplier_list, safe=False, status=200)


class EditSuppliersView(APIView):
    permission_classes = (IsAuthenticated,  IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            user = request.user
            data = json.loads(request.body)

            # Get supplier using id
            supplier_id = data.get('id')
            supplier = Suppliers.objects.get(id=supplier_id, company=user.company)
            dirty_fields = supplier.get_dirty_fields()

            # Update supplier fields
            for field in ['name', 'contact_name', 'contact_no', 'tax_code', 'mail', 'adress', 'explanation']:
                value = data.get(field)
                if value is not None and value != '':
                    setattr(supplier, field, value)
                else:
                    return JsonResponse({'error': _("The field '%s' cannot be empty.") % field}, status=400)

            # Check if any fields have been changed
            if supplier.is_dirty():
                dirty_fields = supplier.get_dirty_fields()

                # Check if the tax_code is being updated and if it's unique within the same company
                if 'tax_code' in dirty_fields:
                    new_tax_code = getattr(supplier, 'tax_code')
                    if Suppliers.objects.filter(tax_code=new_tax_code, company=user.company).exclude(id=supplier_id).exists():
                        return JsonResponse({'error': _("The Tax Code '%s' already exists in the database for another company.") % new_tax_code}, status=400)

            supplier.save()
            return JsonResponse({'message': _("Your changes have been successfully saved.")}, status=200)

        except Suppliers.DoesNotExist:
            return JsonResponse({'error': _("Supplier not found.")}, status=400)

        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class DeleteSupplierView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            id = data.get('id')
            Suppliers.objects.get(id=id).delete()
        except ProtectedError:
            return JsonResponse({
                'error': _("This Supplier cannot be deleted because it is being used elsewhere in the application.")
            }, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        return JsonResponse({'message': _("Supplier object has been successfully deleted")}, status=200)

# endregion

# region Consumers

class AddConsumersView(APIView):
    permission_classes = (IsAuthenticated,  IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            user = request.user
            data = json.loads(request.body)

            tax_code = data.get('tax_code')
            if not tax_code:
                return JsonResponse({'error': _("Tax Code cannot be empty!")}, status=400)
            elif Consumers.objects.filter(tax_code=tax_code, company=user.company).exists():
                error_message = _("The Tax Code '%s' already exists in the database for this company.") % tax_code
                return JsonResponse({'error': error_message}, status=400)

            # List of mandatory fields
            mandatory_fields = ['name', 'contact_name', 'contact_no']

            # Initialize a dictionary to store new supplier data
            new_consumer_data = {}

            for field in ['name', 'contact_name', 'contact_no', 'mail', 'address', 'explanation']:
                value = data.get(field)

                # If the value is not None and not an empty string, save it
                if value is not None and value != '':
                    new_consumer_data[field] = value

                # If the value is None or an empty string, and the field is mandatory, raise an error
                elif field in mandatory_fields:
                    error_message = _("The field '{%s}' cannot be empty.") % field
                    return JsonResponse({'error': error_message}, status=400)

            new_consumer = Consumers.objects.create(tax_code=tax_code, company=user.company, **new_consumer_data)

            message = _("New consumer '{%s}' has been successfully created.") % new_consumer.name
            return JsonResponse({'message': message}, status=201)

        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

class ConsumersView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def get(self, request, *args, **kwargs):
        company = request.user.company
        consumers = Consumers.objects.filter(company=company)
        consumer_list = [[c.id, c.tax_code, c.name, c.contact_name, c.contact_no, c.mail, c.adress, c.explanation] for c in consumers]
        return JsonResponse(consumer_list, safe=False, status=200)

class EditConsumersView(APIView):
    permission_classes = (IsAuthenticated,  IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            user = request.user
            data = json.loads(request.body)
            print(data)

            # Get consumer using id
            consumer_id = data.get('id')
            consumer = Consumers.objects.get(id=consumer_id, company=user.company)

            # Update consumer fields
            for field in ['name', 'contact_name', 'contact_no', 'tax_code', 'mail', 'adress', 'explanation']:
                value = data.get(field)
                if value is not None and value != '':
                    setattr(consumer, field, value)
                else:
                    return JsonResponse({'error': _("The field '%s' cannot be empty.") % field}, status=400)

            # Check if any fields have been changed
            if consumer.is_dirty():
                dirty_fields = consumer.get_dirty_fields()
                print(dirty_fields)

                # Check if the tax_code is being updated and if it's unique within the same company
                if 'tax_code' in dirty_fields:
                    print("omer")
                    new_tax_code = getattr(consumer, 'tax_code')
                    if Consumers.objects.filter(tax_code=new_tax_code, company=user.company).exclude(id=consumer_id).exists():
                        return JsonResponse({'error': _("The Tax Code '%s' already exists in the database for another company.") % new_tax_code}, status=400)

            consumer.save()
            return JsonResponse({'message': _("Your changes have been successfully saved.")}, status=200)

        except Consumers.DoesNotExist:
            return JsonResponse({'error': _("Consumer not found.")}, status=400)

        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class DeleteConsumerView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            id = data.get('id')
            Consumers.objects.get(id=id).delete()
        except ProtectedError:
            return JsonResponse({
                'error': _("This Consumer cannot be deleted because it is being used elsewhere in the application.")
            }, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        return JsonResponse({'message': _("Consumer object has been successfully deleted")}, status=200)

# endregion


# region ProductGroups

class CreateProductGroupView(APIView):
    permission_classes = (IsAuthenticated,  IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        group_name = request.data.get('group_name')

         # Check if group name already exists
        if ProductGroups.objects.filter(group_name=group_name, company=request.user.company).exists():
            return JsonResponse({'error': _('A product group with this name already exists.')}, status=400)

        # Retry the operation up to 5 times
        for i in range(5):
            try:
                with transaction.atomic():
                    # Get the current maximum group code
                    max_group_code = ProductGroups.objects.filter(company=request.user.company).aggregate(Max('group_code'))['group_code__max']

                    if max_group_code is None:
                        # If no group codes exist, start at 1
                        new_group_code = 1
                    else:
                        # Else, increment the highest group code by 1
                        new_group_code = int(max_group_code) + 1

                    # Pad the group code with leading zeros to be 3 digits long
                    padded_group_code = str(new_group_code).zfill(3)

                    # Create the new product group
                    product_group = ProductGroups(group_code=padded_group_code, group_name=group_name, company=request.user.company)
                    product_group.save()

                # If we reach this point, the operation was successful
                return JsonResponse({'message': _('Product group created successfully')}, status=200)

            except IntegrityError:
                # If an IntegrityError is raised, this means another transaction created a product group with the same code.
                # We will retry the operation.
                pass

        # If we reach this point, we have failed to create a product group after several attempts.
        return JsonResponse({'error': _('Could not create product group. Please try again.')}, status=500)

class ProductGroupsView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def get(self, request):
        product_groups = ProductGroups.objects.filter(company=request.user.company)

        # Create a list of lists where each sublist is [group_code, group_name]
        product_groups_list = [[group.group_code, group.group_name] for group in product_groups]
        return JsonResponse(product_groups_list, safe= False)

class EditProductGroupView(APIView):
    permission_classes = (IsAuthenticated,  IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        group_code = request.data.get('group_code')
        new_group_name = request.data.get('new_group_name')

        if not group_code or not new_group_name:
            return JsonResponse({'error': _('Both group_code and new_group_name must be provided')}, status=400)

        try:
            product_group = ProductGroups.objects.get(group_code=group_code, company=request.user.company)
        except ProductGroups.DoesNotExist:
            return JsonResponse({'error': _('Product group not found')}, status=404)

        product_group.group_name = new_group_name
        product_group.save()

        return JsonResponse({'message': _('Product group updated successfully')})
    
#! Bu view kontrol edilmeli, tam olarak düzgün çalışmıyor olabilir.
class DeleteProductGroupView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        group_code = request.data.get('group_code')

        if not group_code:
            return JsonResponse({'error': _('group_code must be provided')}, status=400)

        try:
            product_group = ProductGroups.objects.get(group_code=group_code, company=request.user.company)
        except ProductGroups.DoesNotExist:
            return JsonResponse({'error': _('Product group not found')}, status=404)
        try:
            product_group.delete()
        
            #! Product grup kodlarını güncelleştirmek işlerde tutarsızlıklara yol açabilir.
            # Decrement group_code of subsequent product groups
            ProductGroups.objects.filter(group_code__gt=group_code, company=request.user.company).update(group_code=F('group_code') - 1)
            return JsonResponse({'message': _('Product group deleted successfully')})
        except ProtectedError:
            return JsonResponse({
                'error': _("This product group cannot be deleted because it is being used elsewhere in the application.")
            }, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

# endregion

# region ProductSubgroups

class CreateProductSubgroupView(APIView):
    permission_classes = (IsAuthenticated,  IsSuperStaffOrStockStaff )
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        
        subgroup_name = request.data.get('subgroup_name')
        group_code = request.data.get('group_code')

        if ProductSubgroups.objects.filter(subgroup_name=subgroup_name, group__group_code=group_code, group__company=request.user.company).exists():
            return JsonResponse({'error': _('A product subgroup with this name already exists.')}, status=400)

        # Retry the operation up to 5 times
        for i in range(5):
            try:
                with transaction.atomic():
                    group = ProductGroups.objects.get(group_code=group_code, company=request.user.company)
                    max_subgroup_code = ProductSubgroups.objects.filter(group=group).aggregate(Max('subgroup_code'))['subgroup_code__max']

                    if max_subgroup_code is None:
                        new_subgroup_code = 1
                    else:
                        new_subgroup_code = max_subgroup_code + 1

                    padded_subgroup_code = str(new_subgroup_code).zfill(3)

                    product_subgroup = ProductSubgroups(subgroup_code=padded_subgroup_code, subgroup_name=subgroup_name, group=group)
                    product_subgroup.save()

                return JsonResponse({'message': _('Product subgroup created successfully')})

            except IntegrityError:
                pass

        return JsonResponse({'error': _('Could not create product subgroup. Please try again.')}, status=500)


class ProductSubgroupsView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request):
        group_code = request.data.get('group_code')
        group = ProductGroups.objects.get(group_code=group_code, company=request.user.company)
        product_subgroups = ProductSubgroups.objects.filter(group=group)

        product_subgroups_list = [[subgroup.subgroup_code, subgroup.subgroup_name] for subgroup in product_subgroups]

        return JsonResponse(product_subgroups_list, safe=False)


class EditProductSubgroupView(APIView):
    permission_classes = (IsAuthenticated,  IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        print(request.data)
        subgroup_code = request.data.get('subgroup_code')
        new_subgroup_name = request.data.get('new_subgroup_name')
        group_code = request.data.get('group_code')

        subgroup = ProductSubgroups.objects.get(subgroup_code=subgroup_code,group__group_code=group_code, group__company=request.user.company)
        if not subgroup_code or not new_subgroup_name:
            return JsonResponse({'error': _('Both subgroup_code and new_subgroup_name must be provided')}, status=400)

        if ProductSubgroups.objects.filter(subgroup_name=new_subgroup_name, group__group_code=group_code, group__company=request.user.company).exists():
            return JsonResponse({'error': _('A product subgroup with this name already exists.')}, status=400)
        try:
            product_subgroup = ProductSubgroups.objects.get(subgroup_code=subgroup_code, group__group_code=group_code, group__company=request.user.company)
        except ProductSubgroups.DoesNotExist:
            return JsonResponse({'error': _('Product subgroup not found')}, status=404)

        product_subgroup.subgroup_name = new_subgroup_name
        product_subgroup.save()

        return JsonResponse({'message': _('Product subgroup updated successfully')})


class DeleteProductSubgroupView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        print(request.data)
        subgroup_code = request.data.get('subgroup_code')
        group_code = request.data.get('group_code')

        if not subgroup_code:
            return JsonResponse({'error': _('subgroup_code must be provided')}, status=400)

        try:
            product_subgroup = ProductSubgroups.objects.get(subgroup_code=subgroup_code, group__group_code=group_code, group__company=request.user.company)
        except ProductSubgroups.DoesNotExist:
            return JsonResponse({'error': _('Product subgroup not found')}, status=404)

        try:
            product_subgroup.delete()

            #! subgrup kodu güncellemek ilerde tutarsızlıklara yol açabilir.
            ProductSubgroups.objects.filter(subgroup_code__gt=subgroup_code, group=product_subgroup.group).update(subgroup_code=F('subgroup_code') - 1)

            return JsonResponse({'message': _('Product subgroup deleted successfully')})
        except ProtectedError:
            return JsonResponse({
                'error': _("This product subgroup group cannot be deleted because it is being used elsewhere in the application.")
            }, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

# endregion

# region ProductInflow

class AddProductInflowView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)
        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            # Bill Level Information
            print(request.POST)
            date = request.POST.get('date')
            bill_number = request.POST.get('bill_number')
            provider_company_tax_code = request.POST.get('provider_company_tax_code')
            receiver_company_tax_code = request.POST.get('receiver_company_tax_code')
            company = request.user.company
            project = request.user.current_project

            # Date Parsing
            from datetime import datetime
            date = datetime.strptime(date, '%Y-%m-%d').date()

            # Fetch the companies
            supplier_company = Suppliers.objects.get(tax_code=provider_company_tax_code, company=company)
            receiver_company = Consumers.objects.get(tax_code=receiver_company_tax_code, company=company)

            # Create the ProductInflow object (represents the bill)
            product_inflow = ProductInflow.objects.create(
                date=date,
                bill_number=bill_number,
                supplier_company=supplier_company,
                receiver_company=receiver_company,
                company=company,
                project=project
            )

            # Loop through the list of products
            products = json.loads(request.POST.get('products'))  # Assuming products are sent as a list of dictionaries
            print(products)
            for product_data in products:
                
                product_code = product_data.get('product_code')
                barcode = product_data.get('barcode')
                status = product_data.get('status')
                place_of_use = product_data.get('place_of_use')
                amount = product_data.get('amount')

                product = Products.objects.get(product_code=product_code, company=company)

                # Create the ProductInflowItem (represents an individual product in the bill)
                ProductInflowItem.objects.create(
                    inflow=product_inflow,
                    product=product,
                    barcode=barcode,
                    status=status,
                    place_of_use=place_of_use,
                    amount=amount
                )

            # Process the uploaded images for the bill
            images = request.FILES.getlist('images')
            for image in images:
                # Check that it's an image
                if not isinstance(image, InMemoryUploadedFile) or image.content_type not in ['image/png', 'image/jpeg']:
                    return JsonResponse({'error': _('Invalid file type. Only PNG and JPEG are allowed.')}, status=400)

                # Check file size (max 2 MB)
                if image.size > 2 * 1024 * 1024:
                    return JsonResponse({'error': _('The image file is too large (max 2 MB).')}, status=400)

                # Use Django's built-in validation
                try:
                    img = Image.open(image)
                    img.verify()
                    ProductInflowImage.objects.create(product_inflow=product_inflow, image=image)
                except (IOError, SyntaxError):
                    return JsonResponse({'error': _('One or more uploaded files are not valid images.')}, status=400)

            return JsonResponse({'message': _('ProductInflow created successfully')}, status=201)

        except Products.DoesNotExist:
            return JsonResponse({'error': _('Product with the provided product code does not exist.')}, status=400)
        except Suppliers.DoesNotExist:
            return JsonResponse({'error': _('Supplier with the provided tax code does not exist.')}, status=400)
        except Consumers.DoesNotExist:
            return JsonResponse({'error': _('Consumer with the provided tax code does not exist.')}, status=400)
        except Company.DoesNotExist:
            return JsonResponse({'error': _('Company does not exist.')}, status=400)
        except Project.DoesNotExist:
            return JsonResponse({'error': _('Project does not exist.')}, status=400)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)

class AddProductInflowItemView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)
        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            # Get the ProductInflow object for which we are adding an item
            print(request.data)
            inflow_id = request.data.get('inflow_id')
            product_inflow = ProductInflow.objects.get(id=inflow_id)

            # Fetch the details of the item
            product_code = request.data.get('product_code')
            barcode = request.data.get('barcode')
            status = request.data.get('status')
            place_of_use = request.data.get('place_of_use')
            amount = request.data.get('amount')

            # Get the associated product using the product_code
            product = Products.objects.get(product_code=product_code, company=request.user.company)

            # Create the ProductInflowItem (represents an individual product in the bill)
            ProductInflowItem.objects.create(
                inflow=product_inflow,
                product=product,
                barcode=barcode,
                status=status,
                place_of_use=place_of_use,
                amount=amount
            )

            return JsonResponse({'message': _('ProductInflowItem added successfully')}, status=201)

        except ProductInflow.DoesNotExist:
            return JsonResponse({'error': _('ProductInflow with the provided ID does not exist.')}, status=400)
        except Products.DoesNotExist:
            return JsonResponse({'error': _('Product with the provided product code does not exist.')}, status=400)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)

class ProductInflowView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)
        return super().handle_exception(exc)

    def get(self, request, *args, **kwargs):
        try:
            product_inflows = ProductInflow.objects.filter(
                company=request.user.company,
                project=request.user.current_project
            ).prefetch_related(
                'images',
                'items',
                'items__product',
                'items__product__group',
                'items__product__subgroup'
            ).all()

            product_inflow_list = []
            for pf in product_inflows:
                items = [
                    {
                        'id': item.id,
                        'product_code': item.product.product_code,
                        'barcode': item.barcode,
                        'status': item.status,
                        'place_of_use': item.place_of_use,
                        'group_name': item.product.group.group_name,
                        'subgroup_name': item.product.subgroup.subgroup_name,
                        'brand': item.product.brand,
                        'serial_number': item.product.serial_number,
                        'model': item.product.model,
                        'description': item.product.description,
                        'unit': item.product.unit,
                        'amount': item.amount
                    } for item in pf.items.all()
                ]

                inflow_data = {
                    'id': pf.id,
                    'bill_number': pf.bill_number,
                    'date': pf.date,
                    'supplier_company_tax_code': pf.supplier_company.tax_code,
                    'supplier_company_name': pf.supplier_company.name,
                    'receiver_company_tax_code': pf.receiver_company.tax_code,
                    'receiver_company_name': pf.receiver_company.name,
                    'items': items,
                    'images': [settings.MEDIA_URL + str(image.image) for image in pf.images.all()]
                }
                product_inflow_list.append(inflow_data)

            return JsonResponse(product_inflow_list, safe=False, status=200)

        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)





class EditProductInflowView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)
        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            data = request.data
            inflow_id = data.get('inflow_id')

            # Fetch the inflow 
            inflow = ProductInflow.objects.get(id=inflow_id)

            # Fetch supplier and receiver companies
            supplier_company = Suppliers.objects.get(tax_code=data['supplier_tax_code'], company=request.user.company)
            receiver_company = Consumers.objects.get(tax_code=data['receiver_tax_code'], company=request.user.company)

            # Update the inflow with new details
            inflow.date = data['date']
            inflow.bill_number = data['bill_number']
            inflow.supplier_company = supplier_company
            inflow.receiver_company = receiver_company
            inflow.save()

            return JsonResponse({'message': _("Product Inflow changes have been successfully saved.")}, status=200)

        except ProductInflow.DoesNotExist:
            return JsonResponse({'error': _("Product Inflow not found.")}, status=400)
        except Suppliers.DoesNotExist:
            return JsonResponse({'error': _("Supplier company not found.")}, status=400)
        except Consumers.DoesNotExist:
            return JsonResponse({'error': _("Receiver company not found.")}, status=400)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)

class EditProductInflowItemView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)
        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            data = request.data
            item_id = data.get('item_id')

            # Get the inflow item
            item = ProductInflowItem.objects.get(id=item_id)
            
            # Update the item with new details
            product = Products.objects.get(product_code=data['product_code'], company=request.user.company)
            item.product = product
            item.barcode = data['barcode']
            item.status = data['status']
            item.place_of_use = data['place_of_use']
            item.amount = data['amount']
            item.save()

            return JsonResponse({'message': _("Product Inflow Item changes have been successfully saved.")}, status=200)

        except ProductInflowItem.DoesNotExist:
            return JsonResponse({'error': _("Product Inflow Item not found.")}, status=400)
        except Products.DoesNotExist:
            return JsonResponse({'error': _("Product not found.")}, status=400)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)




#! Ambar girişi silineceği zaman kullanıcıya silmeden önce, muhasebe kaydının da silineceği ile alakalı uyarı verilmeli. 
class DeleteProductInflowView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            print(request.data)
            id = request.data.get('id')
            product_inflow = ProductInflow.objects.get(id=id)

            product_inflow.delete()

        except ProtectedError:
            return JsonResponse({
                'error': _("This product inflow cannot be deleted because it is being used elsewhere in the application.")
            }, status=400)
        except ProductInflow.DoesNotExist:
            return JsonResponse({'error': _('Product inflow object not found.')}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

        return JsonResponse({'message': _('Product inflow object has been successfully deleted.')}, status=200)

class DeleteProductInflowItemView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)
        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            item_id = request.data.get('item_id')
            product_inflow_item = ProductInflowItem.objects.get(id=item_id)
            
            product_inflow_item.delete()

            return JsonResponse({'message': _('ProductInflowItem deleted successfully')}, status=200)

        except ProductInflowItem.DoesNotExist:
            return JsonResponse({'error': _('ProductInflowItem with the provided ID does not exist.')}, status=400)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)

class AddImageView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            # Get the ProductInflow object using its id
            content_type = request.data.get('content_type')
            id = request.data.get('id')
            if content_type == "product_inflow":
                product_inflow = ProductInflow.objects.get(id=id)
            elif content_type == "product_outflow":
                product_outflow = ProductOutflow.objects.get(id=id)

            # Fetch the uploaded images
            images = request.FILES.getlist('images')
            for image in images:
                # Validate the image (type, size, and content)
                if not isinstance(image, InMemoryUploadedFile) or image.content_type not in ['image/png', 'image/jpeg']:
                    return JsonResponse({'error': _('Invalid file type. Only PNG and JPEG are allowed.')}, status=400)

                if image.size > 2 * 1024 * 1024:
                    return JsonResponse({'error': _('The image file is too large (max 2 MB).')}, status=400)

                try:
                    img = Image.open(image)
                    img.verify()
                    if content_type == "product_inflow":
                        ProductInflowImage.objects.create(product_inflow=product_inflow, image=image)
                    elif content_type == "product_outflow":
                        ProductOutflowImage.objects.create(product_outflow=product_outflow, image=image)
                except (IOError, SyntaxError, ValidationError):
                    return JsonResponse({'error': _('One or more uploaded files are not valid images.')}, status=400)

            return JsonResponse({'message': _('Images added successfully to ProductInflow.')}, status=201)

        except ProductInflow.DoesNotExist:
            return JsonResponse({'error': _('ProductInflow with the provided id does not exist.')}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

class DeleteImageView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            # Get content_type and id from the request
            print( request.data)
            content_type = request.data.get('content_type')
            id = request.data.get('id')
            image_file = request.data.get('image')  # Get the image filename you want to delete

            # Based on content_type, fetch the related image model
            if content_type == "product_inflow":
                product_inflow = ProductInflow.objects.get(id=id)
                product_inflow_image = ProductInflowImage.objects.get(product_inflow=product_inflow, image="product_images/" + image_file)
                product_inflow_image.image.delete(save=True)
                product_inflow_image.delete()
            elif content_type == "product_outflow":
                product_outflow = ProductOutflow.objects.get(id=id)
                product_outflow_image = ProductOutflowImage.objects.get(product_outflow=product_outflow, image="product_images/" + image_file)
                product_outflow_image.image.delete(save=True)
                product_outflow_image.delete()
            else:
                return JsonResponse({'error': _('Invalid content_type provided.')}, status=400)

            return JsonResponse({'message': _('Image deleted successfully.')}, status=200)

        except ProductInflowImage.DoesNotExist:
            print(ProductInflowImage.objects.all().values())
            traceback.print_exc()
            return JsonResponse({'error': _('Image associated with the given ProductInflow ID does not exist or filename mismatch.')}, status=400)
        except ProductOutflowImage.DoesNotExist:
            return JsonResponse({'error': _('Image associated with the given ProductOutflow ID does not exist or filename mismatch.')}, status=400)
        except Exception as e:
            traceback.print_exc()

            return JsonResponse({'error': str(e)}, status=500)

# endregion

# region ProductOutflow

class AddProductOutflowView(APIView):
    permission_classes = (IsAuthenticated,  IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            product_code = request.data.get('product_code')
            date = request.data.get('date')
            barcode = request.data.get('barcode')
            provider_company_tax_code = request.data.get('provider_company_tax_code')
            receiver_company_tax_code = request.data.get('receiver_company_tax_code')
            status = request.data.get('status')
            place_of_use = request.data.get('place_of_use')
            amount = request.data.get('amount')
            # Check if the date is in correct format
            try:
                from datetime import datetime
                date = datetime.strptime(date, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'error': _('Invalid date format. Use YYYY-MM-DD.')}, status=400)

            # Fetch the product using the product code
            product = Products.objects.get(product_code=product_code, company=request.user.company)
            supplier_company = Suppliers.objects.get(tax_code=provider_company_tax_code, company=request.user.company)
            receiver_company = Consumers.objects.get(tax_code=receiver_company_tax_code, company=request.user.company)
            # Create the ProductOutflow object
            product_outflow = ProductOutflow.objects.create(
                product=product,
                date=date,
                barcode=barcode,
                supplier_company=supplier_company,
                receiver_company=receiver_company,
                status=status,
                place_of_use=place_of_use,
                amount=amount,
                company=request.user.company,
                project=request.user.current_project
            )
            # Add project to product, consumer, and supplier projects fields
            product.projects.add(request.user.current_project)
            supplier_company.projects.add(request.user.current_project)
            receiver_company.projects.add(request.user.current_project)

            # Get the uploaded images
            images = request.FILES.getlist('images')
            
            for image in images:
                # Check that it's an image
                if not isinstance(image, InMemoryUploadedFile) or image.content_type not in ['image/png', 'image/jpeg']:
                    return JsonResponse({'error': _('Invalid file type. Only PNG and JPEG are allowed.')}, status=400)

                # Check file size (max 2 MB)
                if image.size > 2 * 1024 * 1024:
                    return JsonResponse({'error': _('The image file is too large (max 2 MB).')}, status=400)

                # Use Django's built-in validation
                try:
        # Open the image file
                    img = Image.open(image)

                    # Check if the file is an image
                    img.verify()

                    # The file is an image, save it to your model
                    ProductOutflowImage.objects.create(product_outflow=product_outflow, image=image)
                except (IOError, SyntaxError):
                    # The file is not an image, handle the exception
                    return JsonResponse({'error': _('One or more uploaded files are not valid images.')}, status=400)

            return JsonResponse({'message': _('ProductOutflow created successfully')}, status=201)
        except Products.DoesNotExist:
            return JsonResponse({'error': _('Product with the provided product code does not exist or doesnt belong to your company.')}, status=400)
        except Consumers.DoesNotExist:
            return JsonResponse({'error': _('Consumer with the provided tax code does not exist or doesnt belong to your company.')}, status=400)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)

class ProductOutflowView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def get(self, request, *args, **kwargs):
        product_outflows = ProductOutflow.objects.filter(company=request.user.company, project=request.user.current_project).select_related('product').all()

        product_outflow_list = [
            [
                pf.id, 
                pf.date,
                pf.product.product_code, 
                pf.barcode, 
                pf.supplier_company.tax_code, 
                pf.supplier_company.name, 
                pf.receiver_company.tax_code, 
                pf.receiver_company.name, 
                pf.status, 
                pf.place_of_use,
                pf.product.group.group_name,
                pf.product.subgroup.subgroup_name,
                pf.product.brand,
                pf.product.serial_number,
                pf.product.model,
                pf.product.description,
                pf.product.unit,
                pf.amount,
                [settings.MEDIA_URL + str(image.image) for image in pf.images.all()]  # URLs of all images related to this ProductInflow

                #? [image.image.url for image in pf.images.all()]  # URLs of all images related to this ProductOutflow
            ] 
            for pf in product_outflows
        ]

        return JsonResponse(product_outflow_list, safe=False, status=200)

class EditProductOutflowView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)
        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            data = request.data
            outflow_id = data.get('old_id')
            product_outflow = ProductOutflow.objects.get(id=outflow_id, company=request.user.company, project=request.user.current_project)

            # Update product_outflow fields
            if 'product_code' in data:
                product_outflow.product = Products.objects.get(product_code=data['product_code'], company=request.user.company)
            if 'provider_company_tax_code' in data:
                product_outflow.supplier_company = Consumers.objects.get(tax_code=data['provider_company_tax_code'], company=request.user.company)
            if 'receiver_company_tax_code' in data:
                product_outflow.receiver_company = Consumers.objects.get(tax_code=data['receiver_company_tax_code'], company=request.user.company)

            for field in ['date', 'status', 'place_of_use', 'amount', 'barcode']:
                if field in data:
                    setattr(product_outflow, field, data[field])

            product_outflow.save()

            return JsonResponse({'message': _("Your changes have been successfully saved.")}, status=200)

        except ProductOutflow.DoesNotExist:
            return JsonResponse({'error': _("Product not found.")}, status=400)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)



class DeleteProductOutflowView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            id = request.data.get('id')
            product_outflow = ProductOutflow.objects.get(id=id)

            # No need to check for associated accounting objects as there shouldn't be any

            product_outflow.delete()

        except ProtectedError:
            return JsonResponse({
                'error': _("This product outflow cannot be deleted because it is being used elsewhere in the application.")
            }, status=400)
        except ProductOutflow.DoesNotExist:
            return JsonResponse({'error': _("Product outflow object not found.")}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        return JsonResponse({'message': _("Product outflow object has been successfully deleted.")}, status=200)

class CreateProductOutflowReceiptView(APIView):
    permission_classes = (IsAuthenticated,  IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)
    def post(self, request, *args, **kwargs):
        try:
            id = request.data.get('id')
            product_outflow = ProductOutflow.objects.get(id=id)  # Get the ProductOutflow object
            product_code = product_outflow.product.product_code
            brand = product_outflow.product.brand
            model = product_outflow.product.model
            description = product_outflow.product.description
            product_info = f"{brand} {model} {description}"
            items = [(product_code, product_info, product_outflow.amount, product_outflow.product.unit)]  # Format the ProductOutflow object into 'items' required by 'create_receipt_pdf'
            # For example:
            # items = [[product_outflow.product_code, product_outflow.product_name, product_outflow.quantity, product_outflow.unit]]
            title = _("Ambar Cikis Fisi")
            logo_path = settings.LOGO_PATH  # Update this to the path where your logo is store

            sequence_number_obj = SequenceNumber.objects.first()  # Update this as per your requirement
            # Check if sequence_number_obj is None and create a new instance if necessary
            if not sequence_number_obj:
                sequence_number_obj = SequenceNumber()
                sequence_number_obj.save()

            sequence_number = sequence_number_obj.number
            sequence_number_obj.number += 1
            sequence_number_obj.save()
            # Create a temporary file for the PDF
            fd, temp_pdf_filename = tempfile.mkstemp()
            os.close(fd)
            # Create the PDF
            create_receipt_pdf(temp_pdf_filename, title, items, logo_path, product_code, sequence_number)
            # Generate serial number
            product_code_str = product_code.replace(".", "")
            date_str_serial = datetime.datetime.now().strftime("%d%m%Y")
            serial_number = f"{product_code_str}-{date_str_serial}-{sequence_number:04}"
            # Set filename to be serial_number
            filename = f'{serial_number}.pdf'

            with open(temp_pdf_filename, 'rb') as f:
                pdf = f.read()
                base64_content = base64.b64encode(pdf).decode()

            # Remove the temporary file
            os.remove(temp_pdf_filename)

            return JsonResponse({'filename': filename, 'content': base64_content})

        except ProductOutflow.DoesNotExist:
            return JsonResponse({'error': "Product Outflow not found."}, status=400)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)
        

# endregion

# region Stock

# @receiver(pre_save, sender=Products)
# def create_stock(sender, instance, created, **kwargs):
#     if created:
#         Stock.objects.create(
#             product=instance,
#             warehouse=None,
#             inflow=0,
#             outflow=0,
#             stock=0,
#             company=instance.company,
#             project=instance.project
#         )

#!!!!!!
# @receiver(post_save, sender=ProductInflowItem)
# def update_stock_inflow(sender, instance, created, **kwargs):
#     if created:  # Only handle inflow for newly created items
#         try:
#             stock = Stock.objects.get(product=instance.product, company=instance.inflow.company, project=instance.inflow.project)
#             stock.inflow += float(instance.amount)
#             stock.stock = (stock.inflow or 0) - (stock.outflow or 0)
#             stock.save()

#             instance.inflow.supplier_company.products.add(instance.product)
#             instance.inflow.receiver_company.products.add(instance.product)
#         except Stock.DoesNotExist:
#             return JsonResponse({'error': _("Product could not be found on Stock.")}, status=400)


# @receiver(pre_delete, sender=ProductInflowItem)
# def update_stock_inflow_on_delete(sender, instance, **kwargs):
#     try:
#         stock = Stock.objects.get(product=instance.product, company=instance.inflow.company, project=instance.inflow.project)
#         stock.inflow -= instance.amount
#         stock.stock = (stock.inflow or 0) - (stock.outflow or 0)
#         stock.save()

#         # If product was added to the supplier, remove it
#         instance.inflow.supplier_company.products.remove(instance.product)
#         instance.inflow.receiver_company.products.remove(instance.product)
#     except Stock.DoesNotExist:
#         return JsonResponse({'error': _("Product could not be found on Stock.")}, status=400)




# @receiver(post_save, sender=ProductOutflow)
# def update_stock_outflow(sender, instance, created, **kwargs):
#     with transaction.atomic():
#         dirty_fields = instance.get_dirty_fields()

#         old_product_code = dirty_fields.get('product__product_code')
#         if old_product_code:
#             try:
#                 old_product_stock = Stock.objects.get(product__product_code=old_product_code, company=instance.company, project=instance.project)
#                 old_product_stock.outflow -= instance.amount
#                 old_product_stock.stock = (old_product_stock.inflow or 0) - (old_product_stock.outflow or 0)
#                 if old_product_stock.stock < 0:
#                     return JsonResponse({'error': _("You do not have enough products in your warehouse.")}, status=400)
#                 old_product_stock.save()
#             except Stock.DoesNotExist:
#                 return JsonResponse({'error': _("Product could not be found on Stock.")}, status=400)

#         try:
#             stock = Stock.objects.get(product__product_code=instance.product.product_code, company=instance.company, project=instance.project)
#         except Stock.DoesNotExist:
#             return

#         if created:
#             instance.receiver_company.products.add(instance.product)
#             stock.outflow += float(instance.amount)
#         else:
#             old_amount = dirty_fields.get('amount')
#             if old_amount is not None:
#                 stock.outflow = (stock.outflow or 0) - old_amount + instance.amount

#         stock.stock = (stock.inflow or 0) - (stock.outflow or 0)
#         if stock.stock < 0:
#             raise ValidationError(_('You do not have enough products in your warehouse.'))
#         stock.save()



# @receiver(pre_delete, sender=ProductOutflow)
# def update_stock_outflow_on_delete(sender, instance, **kwargs):
#     try:
#         stock = Stock.objects.get(product=instance.product, company=instance.company, project=instance.project)
#         stock.outflow -= instance.amount
#         stock.stock = (stock.inflow or 0) - (stock.outflow or 0)
#         stock.save()

#         # If product was added to the receiver, remove it
#         instance.receiver_company.products.remove(instance.product)
#     except Stock.DoesNotExist:
#         return JsonResponse({'error': _("Product could not be found on Stock.")}, status=400)




class StockView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def get(self, request, *args, **kwargs):
        stocks = Stock.objects.filter(company=request.user.company, project=request.user.current_project)
        stock_list = [
            [
                stock.id,
                stock.product.product_code, 
                stock.product.group.group_name, 
                stock.product.subgroup.subgroup_name,
                stock.product.brand, 
                stock.product.serial_number, 
                stock.product.model, 
                stock.product.description,
                stock.product.unit, 
                stock.warehouse, 
                stock.inflow, 
                stock.outflow, 
                stock.stock,
                stock.reserve_stock
            ] 
            for stock in stocks
        ]
        print(stock_list)
        return JsonResponse(stock_list, safe=False, status=200)

#! Kullanıcı stock üzerinde edit yapamasın diye bu view kapatıldı.
# class EditStockView(APIView):
#     permission_classes = (IsAuthenticated, IsSuperStaff)
#     authentication_classes = (JWTAuthentication,)

    # def handle_exception(self, exc):
    #     if isinstance(exc, (NotAuthenticated, PermissionDenied)):
    #         return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

    #     return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             data = json.loads(request.body)

#             product_code = data.get('product_code')
#             stock = Stock.objects.get(product__product_code=product_code)

#             # Update stock fields
#             for field in ['new_warehouse', 'new_inflow', 'new_outflow', 'new_stock']:
#                 value = data.get(field)
#                 if value is not None and value != '':
#                     updated_field = field[4:]  # Remove the "new_" prefix
#                     setattr(stock, updated_field, value)
#                 else:
#                     return JsonResponse({'error': f"The field '{field}' cannot be empty."}, status=400)

#             stock.save()
#             return JsonResponse({'message': f"Your changes have been successfully saved."}, status=200)

#         except Stock.DoesNotExist:
#             return JsonResponse({'error': "Stock not found."}, status=400)
#         except Products.DoesNotExist:
#             return JsonResponse({'error': "Product not found."}, status=400)
#         except ValueError as e:
#             return JsonResponse({'error': str(e)}, status=400)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

#! Kullanıcı stock üzerinde delete yapamasın diye bu view kapatıldı.       
# class DeleteStockView(APIView):
#     permission_classes = (IsAuthenticated, IsSuperStaff)
#     authentication_classes = (JWTAuthentication,)

    # def handle_exception(self, exc):
    #     if isinstance(exc, (NotAuthenticated, PermissionDenied)):
    #         return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

    #     return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             product_code = request.POST.get('product_code')
#             Stock.objects.filter(product_code=product_code).delete()
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)
#         return JsonResponse({'message': "Stock object has been successfully deleted"}, status=200)


# endregion

# region Accounting

@receiver(post_save, sender=ProductInflow)
def create_accounting_inflow(sender, instance, created, **kwargs):
    if created:
        # Create AccountingInflow instance when a new ProductInflow instance is created
        AccountingInflow.objects.create(
            product_inflow=instance,
            price_without_tax=0,  # initial values which will be updated by AccountingItem signals
            price_with_tevkifat=0,
            price_total=0
        )

@receiver(post_save, sender=AccountingItem)
def update_accounting_inflow(sender, instance, **kwargs):
    # instance here refers to the AccountingItem that was just saved
    associated_inflow = instance.accounting_inflow
    associated_inflow.save()  # This will trigger the save() method of AccountingInflow and recalculate its totals

@receiver(post_save, sender=ProductInflowItem)
def create_accounting_item(sender, instance, created, **kwargs):
    if created:
        # Get associated AccountingInflow from ProductInflow
        accounting_inflow = AccountingInflow.objects.get(product_inflow=instance.inflow)
        
        # Create AccountingItem instance when a new ProductInflowItem instance is created
        item = AccountingItem.objects.create(
            accounting_inflow=accounting_inflow,
            product_item=instance,
            # ... add other initial fields if necessary ...
            unit_price=0,  # default values which will be updated when the item is saved
            discount_rate=0,
            discount_amount=0,
            tax_rate=0,
            tevkifat_rate=0,
            price_without_tax=0,
            unit_price_without_tax=0,
            price_with_tevkifat=0,
            price_total=0
        )
        
        # Saving the AccountingItem will trigger its save() method
        item.save()

        # After adding a new AccountingItem, update the associated AccountingInflow
        accounting_inflow.save()




class AccountingView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)
        return super().handle_exception(exc)

    def get(self, request, *args, **kwargs):
        try:
            accounting_inflows = AccountingInflow.objects.filter(
                product_inflow__company=request.user.company,
                product_inflow__project=request.user.current_project
            ).prefetch_related(
                'product_inflow__images',
                'items',
                'items__product_item__product',
                'items__product_item__product__group',
                'items__product_item__product__subgroup'
            ).all()

            accounting_inflow_list = []
            for af in accounting_inflows:
                items = [
                    {
                        'id': item.id,
                        'product_code': item.product_item.product.product_code,
                        'barcode': item.product_item.barcode,
                        'status': item.product_item.status,
                        'place_of_use': item.product_item.place_of_use,
                        'group_name': item.product_item.product.group.group_name,
                        'subgroup_name': item.product_item.product.subgroup.subgroup_name,
                        'brand': item.product_item.product.brand,
                        'serial_number': item.product_item.product.serial_number,
                        'model': item.product_item.product.model,
                        'description': item.product_item.product.description,
                        'unit': item.product_item.product.unit,
                        'amount': item.product_item.amount,
                        'unit_price': item.unit_price,
                        'discount_rate': item.discount_rate,
                        'discount_amount' : item.discount_amount,
                        'tax_rate' : item.tax_rate,
                        'tevkifat_rate' : item.tevkifat_rate,
                        'price_without_tax' : item.price_without_tax,
                        'unit_price_without_tax' : item.unit_price_without_tax,
                        'price_with_tevkifat' : item.price_with_tevkifat,
                        'price_total' : item.price_total,
                    } for item in af.items.all()
                ]

                inflow_data = {
                    'id': af.id,
                    'bill_number': af.product_inflow.bill_number,
                    'date': af.product_inflow.date,
                    'supplier_company_tax_code': af.product_inflow.supplier_company.tax_code,
                    'supplier_company_name': af.product_inflow.supplier_company.name,
                    'receiver_company_tax_code': af.product_inflow.receiver_company.tax_code,
                    'receiver_company_name': af.product_inflow.receiver_company.name,
                    'price_without_tax': af.price_without_tax,
                    'price_with_tevkifat': af.price_with_tevkifat,
                    'price_total': af.price_total,
                    'items': items,
                    'images': [settings.MEDIA_URL + str(image.image) for image in af.product_inflow.images.all()]
                }
                accounting_inflow_list.append(inflow_data)
                print(accounting_inflow_list)
            return JsonResponse(accounting_inflow_list, safe=False, status=200)

        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)



#! Total Price'lar edit yapılamamalı. Onlar diğer verilere göre otomatik hesaplanıyor.
class EditAccountingItemView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            data = request.data
            print(data)
            item_id = data.get('item_id')  # Get the ID of the AccountingItem

            # Filter AccountingItem based on user's company, project, and the provided item ID
            accounting_item = AccountingItem.objects.filter(accounting_inflow__product_inflow__company=request.user.company, accounting_inflow__product_inflow__project=request.user.current_project).get(id=item_id)

            # Update accounting item fields
            for field in ['unit_price', 'discount_rate', 'discount_amount', 'tax_rate', 'tevkifat_rate', 'price_without_tax', 'unit_price_without_tax', 'price_with_tevkifat', 'price_total']:
                value = data.get(field)

                if value is not None and value != '':
                    setattr(accounting_item, field, value)
                else:
                    return JsonResponse({'error': f"The field 'new_{field}' cannot be empty."}, status=400)

            accounting_item.save()
            return JsonResponse({'message': _("Your changes have been successfully saved.")}, status=200)

        except AccountingItem.DoesNotExist:
            return JsonResponse({'error': _("Accounting item record not found.")}, status=400)

        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)





#! muhasebenin kayıt silmesi engellendi, kayıtları sadece amar girişi üzerinden silinebilir. Ambar girişinde silinen 
#! ürün otomatik olarak muhasebe kısmından da silinecektir.
# class DeleteAccountingView(APIView):
#     permission_classes = (IsAuthenticated, IsSuperStaff)
#     authentication_classes = (JWTAuthentication,)

    # def handle_exception(self, exc):
    #     if isinstance(exc, (NotAuthenticated, PermissionDenied)):
    #         return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

    #     return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             id = request.POST.get('id')
#             Accounting.objects.filter(id=id).delete()
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)
#         return JsonResponse({'message': "Accounting object has been successfully deleted"}, status=200)

# endregion

# region Search Supplier and Consumer

class SupplierSearchView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaffOrAccountingStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        product_code = data.get('product_code')
        print(product_code)
        if not product_code:
            return JsonResponse({'error': _('Missing product_code parameter')}, status=400)

        suppliers = Suppliers.objects.filter(products__product_code=product_code, company=request.user.company)
        suppliers_data = [{'id': supplier.id, 'name': supplier.name, 'contact_name': supplier.contact_name, 'contact_no': supplier.contact_no} for supplier in suppliers]
        print(suppliers_data)
        return JsonResponse({'suppliers': suppliers_data})

class ConsumerSearchView(APIView):
    permission_classes = (IsAuthenticated,  IsSuperStaffOrAccountingStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        product_code = data.get('product_code')
        if not product_code:
            return JsonResponse({'error': _('Missing product_code parameter')}, status=400)

        consumers = Consumers.objects.filter(products__product_code=product_code, company=request.user.company)
        consumers_data = [{'id': consumer.id, 'name': consumer.name, 'contact_name': consumer.contact_name, 'contact_no': consumer.contact_no } for consumer in consumers]

        return JsonResponse({'consumers': consumers_data})

# endregion

# region Filter and Search Models

class SupplierConsumerProductSearchView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def get(self, request, *args, **kwargs):
        company = request.user.company

        if not company:
            return JsonResponse({'error': _('Company not associated with user.')}, status=400)

        products = Products.objects.filter(company=company)
        product_list = [[p.id, p.product_code, p.description] for p in products]

        suppliers = Suppliers.objects.filter(company=company)
        supplier_list = [[s.id, s.tax_code, s.name] for s in suppliers]

        consumers = Consumers.objects.filter(company=company)
        consumer_list = [[c.id, c.tax_code, c.name] for c in consumers]
        
        return JsonResponse([product_list, supplier_list, consumer_list], safe=False, status=200)


# endregion

# region QTO

class AddQTOView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': "You do not have permission to perform this action."}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            # Assuming all these fields are provided in the POST request
            project = request.user.current_project
            print(request.data.get('building_id'))
            building_id = request.data.get('building_id')
            elevation_or_floor_id = request.data.get('elevation_or_floor_id')
            section_id = request.data.get('section_id')
            place_id = request.data.get('place_id')

            # You'd fetch the instances using the IDs
            building = Building.objects.get(id=building_id, project=project)
            elevation_or_floor = ElevationOrFloor.objects.get(id=elevation_or_floor_id, building=building)
            try:
                section = Section.objects.get(id=section_id, elevation_or_floor=elevation_or_floor)
            except Place.DoesNotExist:
                section = None
            try:
                place = Place.objects.get(id=place_id, section=section)
            except Place.DoesNotExist:
                place = None

            # Rest of the fields
            pose_code = request.data.get('pose_code')
            pose_number = request.data.get('pose_number')
            manufacturing_code = request.data.get('manufacturing_code')
            material = request.data.get('material')
            description = request.data.get('description')
            width = request.data.get('width')
            depth = request.data.get('depth')
            height = request.data.get('height')
            quantity = request.data.get('quantity')
            unit = request.data.get('unit')
            multiplier = request.data.get('multiplier')
            multiplier2 = request.data.get('multiplier2')
            take_out = request.data.get('take_out')
            # ... and so on for other fields

            qto = QuantityTakeOff(
                pose_code=pose_code,
                pose_number=pose_number,
                manufacturing_code=manufacturing_code,
                material=material,
                description=description,
                width=width or 1,
                depth=depth or 1,
                height=height or 1,
                quantity=quantity,
                unit=unit,
                multiplier=multiplier or 1,
                multiplier2=multiplier2 or 1,
                take_out=take_out,
                project=project,
                building=building,
                elevation_or_floor=elevation_or_floor,
                section=section,
                place=place
            )
            qto.save()
            
            return JsonResponse({'message': 'Quantity Takeoff object created successfully'}, status=201)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)

class AddExcelQTOView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)
        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            if 'file' not in request.FILES:
                return JsonResponse({'error': _("No file uploaded")}, status=400)
            
            file = request.FILES['file']
            kind = filetype.guess(file.read())
            
            if kind is None or kind.mime not in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
                return JsonResponse({'error': _("The uploaded file is not a valid Excel file")}, status=400)

            data = pd.read_excel(file)
            if data.empty:
                return JsonResponse({'error': _("The uploaded file is empty")}, status=400)

            count = 0
            for i, row in data.iterrows():

                # Fetching based on hierarchy and names
                try:
                    project = request.user.current_project

                    building = Building.objects.get(name=row["Building Name"], project=project) if "Building Name" in row and pd.notna(row["Building Name"]) else None

                    elevation_or_floor = ElevationOrFloor.objects.get(name=row["Elevation or Floor Name"], building=building) if "Elevation or Floor Name" in row and pd.notna(row["Elevation or Floor Name"]) else None

                    section = Section.objects.get(name=row["Section Name"], elevation_or_floor=elevation_or_floor) if "Section Name" in row and pd.notna(row["Section Name"]) else None

                    place = Place.objects.get(name=row["Place Name"], section=section) if "Place Name" in row and pd.notna(row["Place Name"]) else None

                except Project.DoesNotExist:
                    return JsonResponse({'error': _("No project found with name '%s'") % row['Project Name']}, status=400)
                except Building.DoesNotExist:
                    return JsonResponse({'error': f"No building found with name '{row['Building Name']}' for project '{row['Project Name']}'"}, status=400)
                except ElevationOrFloor.DoesNotExist:
                    return JsonResponse({'error': f"No elevation or floor found with name '{row['Elevation or Floor Name']}' for building '{row['Building Name']}'"}, status=400)
                except Section.DoesNotExist:
                    return JsonResponse({'error': f"No section found with name '{row['Section Name']}' for elevation/floor '{row['Elevation or Floor Name']}'"}, status=400)
                except Place.DoesNotExist:
                    return JsonResponse({'error': f"No place found with name '{row['Place Name']}' for section '{row['Section Name']}'"}, status=400)

                qto = QuantityTakeOff(
                        project=project,
                        building=building,
                        elevation_or_floor=elevation_or_floor,
                        section=section,
                        place=place,
                        pose_code=row["Pose Code"],
                        pose_number=row["Pose Number"],
                        manufacturing_code=row["Manufacturing Code"],
                        material=row["Material"],
                        description=row["Description"] if "Description" in row and pd.notna(row["Description"]) else None,
                        width=row["Width"],
                        depth=row["Depth"],
                        height=row["Height"],
                        quantity=row["Quantity"],
                        unit=row["Unit"],
                        multiplier=row["Multiplier"] if "Multiplier" in row and pd.notna(row["Multiplier"]) else 1,
                        multiplier2=row["Multiplier2"] if "Multiplier2" in row and pd.notna(row["Multiplier2"]) else 1,
                        take_out=row["Take Out"],
                        total=row["Total"]
                    )

                qto.save()
                count += 1

            return JsonResponse({'message': f"{count} QTO entries added successfully"}, status=200)

        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)

class EditQTOView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)  # Assuming you want the same permissions
    authentication_classes = (JWTAuthentication,)  # Assuming you're using JWT for auth

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            
            data = request.data
            print(data)
            old_id = data.get('old_id')
            qto_instance = QuantityTakeOff.objects.filter(project=request.user.current_project).get(id=old_id)
            
            # Delete the old QTO instance
            qto_instance.delete()

            # # Create a new QTO instance
            # project_id = data.get('project_id')
            # project = Project.objects.get(id=project_id) if project_id else None

            building_id = data.get('building_id')
            building = Building.objects.get(id=building_id) if building_id else None

            elevation_or_floor_id = data.get('elevation_or_floor_id')
            elevation_or_floor = ElevationOrFloor.objects.get(id=elevation_or_floor_id) if elevation_or_floor_id else None

            section_id = data.get('section_id')
            section = Section.objects.get(id=section_id) if section_id else None

            place_id = data.get('place_id')
            place = Place.objects.get(id=place_id) if place_id else None

            qto_data = {
                'project': request.user.current_project,
                'building': building,
                'elevation_or_floor': elevation_or_floor,
                'section': section,
                'place': place,
                # ... and so on for the other fields ...
            }

            for field in ['pose_code', 'pose_number', 'manufacturing_code', 'material', 'description', 'width', 'depth', 'height', 'quantity', 'unit', 'multiplier', 'multiplier2', 'take_out', 'total']:
                value = data.get(field)  # Converting to "Pose Code" format
                if value is not None and value != '':
                    qto_data[field] = value

            new_qto_instance = QuantityTakeOff.objects.create(**qto_data)

            return JsonResponse({'message': _("Your changes have been successfully saved.")}, status=200)

        except QuantityTakeOff.DoesNotExist:
            return JsonResponse({'error': _("QTO not found.")}, status=400)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)

class DeleteQTOView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaff)  # Assuming you want the same permissions
    authentication_classes = (JWTAuthentication,)  # Assuming you're using JWT for auth

    def handle_exception(self, exc):
        if isinstance(exc, (NotAuthenticated, PermissionDenied)):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

        return super().handle_exception(exc)

    def post(self, request, *args, **kwargs):
        try:
            qto_id = request.data.get('qto_id')
            
            if not qto_id:
                return JsonResponse({'error': _("QTO ID not provided.")}, status=400)
                
            qto_instance = QuantityTakeOff.objects.filter(project=request.user.current_project).get(id=qto_id)
            
            # Delete the QTO instance
            qto_instance.delete()

            return JsonResponse({'message': _("The QTO has been successfully deleted.")}, status=200)

        except QuantityTakeOff.DoesNotExist:
            return JsonResponse({'error': _("QTO not found.")}, status=400)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)
        

class BuildingsForProjectView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def get(self, request):
        project = request.user.current_project
        buildings = Building.objects.filter(project=project)
        data = [{'id': building.id, 'name': building.name} for building in buildings]
        return JsonResponse(data, safe=False)

class ElevationsForBuildingView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def post(self, request):
        # Extract building_id from request data (body) instead of GET parameters
        building_id = request.data.get('building_id')
        if not building_id:
            return JsonResponse({'error': 'building_id is required'}, status=400)

        elevations = ElevationOrFloor.objects.filter(building_id=building_id)
        data = [{'id': elevation.id, 'name': elevation.name} for elevation in elevations]
        return JsonResponse(data, safe=False)

class SectionsForElevationView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)

    def post(self, request):
        elevation_id = request.data.get('elevation_id')
        sections = Section.objects.filter(elevation_or_floor_id=elevation_id)
        data = [{'id': section.id, 'name': section.name} for section in sections]
        return JsonResponse(data, safe=False)

class PlacesForSectionView(APIView):
    permission_classes = (IsAuthenticated, IsSuperStaffOrStockStaff)
    authentication_classes = (JWTAuthentication,)
    
    def post(self, request):
        section_id = request.data.get('section_id')
        places = Place.objects.filter(section_id=section_id)
        data = [{'id': place.id, 'name': place.name} for place in places]
        return JsonResponse(data, safe=False)

class QTOView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def handle_exception(self, exc):
        if isinstance(exc, NotAuthenticated):
            return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)
        return super().handle_exception(exc)

    def get(self, request, *args, **kwargs):
        try:
            # Assuming the current project is stored in request.user.current_project
            project = request.user.current_project

            # Filter QTO objects by the user's current project
            qtos = QuantityTakeOff.objects.filter(project=project)

            qto_entries=  [[qto.id, 
                            {'id': qto.building.id, 'name': qto.building.name} if qto.building else None,
                            {'id': qto.elevation_or_floor.id, 'name': qto.elevation_or_floor.name} if qto.elevation_or_floor else None, 
                            {'id': qto.section.id, 'name': qto.section.name} if qto.section else None, 
                            {'id': qto.place.id, 'name': qto.place.name} if qto.place else None,
                qto.pose_code, qto.pose_number, qto.manufacturing_code, qto.material, qto.description, qto.width, qto.depth, qto.height, 
                qto.quantity, qto.unit, qto.multiplier, qto.multiplier2, qto.take_out, qto.total] for qto in qtos]
                
            

            return JsonResponse(list(qto_entries), safe=False)
        
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)



# endregion




# # region Customers

# class AddCustomersView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

    # def handle_exception(self, exc):
    #     if isinstance(exc, (NotAuthenticated, PermissionDenied)):
    #         return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

    #     return super().handle_exception(exc)
    
#     def post(self, request, *args, **kwargs):
#         try:
#             if 'file' not in request.FILES:
#                 raise ValidationError("No file uploaded")
#             file = request.FILES['file']
#             kind = filetype.guess(file.read())
#             if kind is None or kind.mime not in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
#                 return JsonResponse({'error': "The uploaded file is not a valid Excel file!"}, status=400)

#             data = pd.read_excel(file)
#             if data.empty:
#                 return JsonResponse({'error': "The uploaded file is empty"}, status=400)
#             count = 0
#             for i, row in data.iterrows():
#                 try:
#                     customer_code = row["Customer Code"]
#                     if Customers.objects.filter(customer_code=customer_code).exists():
#                         continue
#                     count+=1
#                     customer = Customers(customer_code=customer_code, description=row["Description"], quantity=row["Quantity"],
#                                         area_code=row["Area Code"], code=row["Code"], city=row["City"], area=row["Area"])
#                     customer.save()
#                 except KeyError as e:
#                     return JsonResponse({'error': f"Column '{e}' not found in the uploaded file"}, status=400)
#             return JsonResponse({'message': f"{count} Customers data added successfully"}, status=200)
#         except OperationalError as e:
#             return JsonResponse({'error': f"Database error: {str(e)}"}, status=500)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

# class ViewCustomersView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

    # def handle_exception(self, exc):
    #     if isinstance(exc, (NotAuthenticated, PermissionDenied)):
    #         return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

    #     return super().handle_exception(exc)
#     def get(self,request,*args, **kwargs):
#          customers = Customers.objects.values().all()
#          customer_list = [[customer['customer_code'], customer['description'], customer['quantity'],
#                       customer['area_code'], customer['code'], customer['city'], customer['area']]
#                      for customer in customers]
#          return JsonResponse(customer_list,safe=False)

# class DeleteCustomerView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

    # def handle_exception(self, exc):
    #     if isinstance(exc, (NotAuthenticated, PermissionDenied)):
    #         return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

    #     return super().handle_exception(exc)
#     def post(self, request, *args, **kwargs):
#         try:
#             customer_code = request.POST.get('customer_code')
#             Customers.objects.filter(customer_code=customer_code).delete()
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)
#         return JsonResponse({'message': "Customer object has been successfully deleted"}, status=200)


# class EditCustomerView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

    # def handle_exception(self, exc):
    #     if isinstance(exc, (NotAuthenticated, PermissionDenied)):
    #         return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

    #     return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             data = json.loads(request.body)

#             # Check if old_customer_code is provided
#             old_customer_code = data.get('old_customer_code')
#             if not old_customer_code:
#                 return JsonResponse({'error': "Missing required parameter: old_customer_code"}, status=400)

#             # Get the customer object
#             customer = Customers.objects.get(customer_code=old_customer_code)

#             # Check if new_customer_code value is unique and not empty
#             new_customer_code = data.get('new_customer_code')
#             if new_customer_code and new_customer_code != old_customer_code:
#                 if not new_customer_code:
#                     return JsonResponse({'error': "Customer Code cannot be empty!"}, status=400)
#                 if Customers.objects.filter(customer_code=new_customer_code).exists():
#                     return JsonResponse({'error': f"The customer code '{new_customer_code}' already exists in the database."}, status=400)
#                 else:
#                     customer.customer_code = new_customer_code

#             # Update other customer fields
#             try:
#                 for field in ['description', 'quantity', 'area_code', 'code', 'city', 'area']:
#                     value = data.get(f'new_{field}')
#                     if value is not None and value != '':
#                         setattr(customer, field, value)
#                     else: 
#                         return JsonResponse({'error': "One or more data field is empty!"}, status=400)
#             except Exception as e:
#                 return JsonResponse({'error': str(e)}, status=400)


#             customer.save()
#             return JsonResponse({'message': "Your changes have been successfully saved"}, status=200)

#         except Customers.DoesNotExist:
#             return JsonResponse({'error': "Customer not found!"}, status=400)

#         except ValueError as e:
#             return JsonResponse({'error': str(e)}, status=400)

#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

# class ExportCustomersView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

    # def handle_exception(self, exc):
    #     if isinstance(exc, (NotAuthenticated, PermissionDenied)):
    #         return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

    #     return super().handle_exception(exc)

#     def get(self, request, *args, **kwargs):
#         def set_column_widths(worksheet):
#             for column_cells in worksheet.columns:
#                 length = max(len(str(cell.value)) for cell in column_cells)
#                 worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
#         customers = Customers.objects.all().values()
#         # Create a new workbook and add a worksheet
#         jalali_date= current_jalali_date().strftime('%Y-%m-%d')
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = f"Customers {jalali_date}"
#         # Write the header row
#         header = ['Customer Code', 'Description', 'Quantity', 'Area Code', 'Code', 'City', 'Area']
#         for col_num, column_title in enumerate(header, 1):
#             cell = ws.cell(row=1, column=col_num)
#             cell.value = column_title
#             cell.font = openpyxl.styles.Font(bold=True)
#             cell.fill = openpyxl.styles.PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')
#             cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='medium'),
#                                                  bottom=openpyxl.styles.Side(style='medium'),
#                                                  left=openpyxl.styles.Side(style='medium'),
#                                                  right=openpyxl.styles.Side(style='medium'))
#         # Write the data rows
#         for row_num, customer in enumerate(customers, 2):
#             row = [customer['customer_code'], customer['description'], customer['quantity'],
#                    customer['area_code'], customer['code'], customer['city'], customer['area']]
#             for col_num, cell_value in enumerate(row, 1):
#                 cell = ws.cell(row=row_num, column=col_num)
#                 cell.value = cell_value
#         # Apply some styling to the Excel file
#         for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#             for cell in row:
#                 cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'),
#                                                      bottom=openpyxl.styles.Side(style='thin'),
#                                                      left=openpyxl.styles.Side(style='thin'),
#                                                      right=openpyxl.styles.Side(style='thin'))
#                 cell.alignment = openpyxl.styles.Alignment(horizontal='center')
#         # Set the column widths
#         set_column_widths(ws)

#         # Apply auto filter
#         ws.auto_filter.ref = f"A1:G{ws.max_row}"


#         # Set the response headers for an Excel file
#         buffer = BytesIO()
#         wb.save(buffer)
#         buffer.seek(0)
#         content = buffer.read()

#         # Encode the content in base64
#         base64_content = base64.b64encode(content).decode()


#         # Send the content and filename in the JSON response
#         return JsonResponse({'filename': f'customers({jalali_date}).xlsx', 'content': base64_content})


        
    

# # endregion

# # region Sales

# class AddSalesView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             if 'file' not in request.FILES:
#                 return JsonResponse({'error': "No file uploaded"}, status=400)
            
#             file = request.FILES['file']
#             kind = filetype.guess(file.read())
            
#             if kind is None or kind.mime not in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
#                 return JsonResponse({'error': "The uploaded file is not a valid Excel file"}, status=400)

#             data = pd.read_excel(file)
#             if data.empty:
#                 return JsonResponse({'error': "The uploaded file is empty"}, status=400)

#             count = 0
#             for i, row in data.iterrows():
#                 no = row["No"]
#                 if Sales.objects.filter(no=no).exists():
#                     continue
#                 count += 1

#                 # Check required fields
#                 for field in ['Good Code', 'Customer Code', 'Original Output Value', 'Net Sales', 'Saler', 'PSR']:
#                     if not row[field]:
#                         return JsonResponse({'error': f"{field} cannot be empty"}, status=400)

#                 # Check integer fields
#                 for field in ['Good Code', 'Customer Code']:
#                     try:
#                         if not isinstance(int(row[field]), int):
#                             return JsonResponse({'error': f"{field} should be integer"}, status=400)
#                     except Exception:
#                             return JsonResponse({'error': f"{field} should be integer"}, status=400)


#                 # Check valid date format
#                 try:
#                     date = jdatetime.date(int(row["Year"]), int(row["Month"]), int(row["Day"]))
#                 except ValueError:
#                     return JsonResponse({'error': "Date should be in the format of YYYY-MM-DD"}, status=400)
#                 except IndexError as e:
#                     return JsonResponse({'error': "Date should be in the format of YYYY-MM-DD"}, status=400)
#                 except Exception as e:
#                     return JsonResponse({'error': "Date should be in the format of YYYY-MM-DD"}, status=400)

#                 # Check valid psr value
#                 if row['PSR'] not in ['P', 'S', 'R']:
#                     return JsonResponse({'error': "Invalid P-S-R value. Allowed values are 'P', 'S', and 'R'."}, status=400)

#                 # Check for existing good
#                 if not Warehouse.objects.filter(product_code=row['Good Code']).exists():
#                     return JsonResponse({'error': f"No product found with code '{row['Good Code']}' in warehouse"}, status=400)
                    

#                 # Check for existing customer
#                 if not Customers.objects.filter(customer_code=row['Customer Code']).exists():
#                     return JsonResponse({'error': f"No customer found with code '{row['Customer Code']}'"}, status=400)

#                 # Check for existing saler
#                 if not Salers.objects.filter(name=row['Saler']).exists():
#                     return JsonResponse({'error': f"No saler found with name '{row['Saler']}'"}, status=400)

#                 try:
#                     customer = Customers.objects.get(customer_code= row["Customer Code"] )
#                 except Exception as e:
#                     return JsonResponse({'error': "No customer found"}, status=400)
#                 try:
#                     product = Products.objects.get(product_code_ir= row["Good Code"] )
#                 except Exception as e:
#                     return JsonResponse({'error': "No product found"}, status=400)
#                 try:
#                     saler = Salers.objects.get(name= row["Saler"] )
#                 except Exception as e:
                    
#                     return JsonResponse({'error': "No saler found"}, status=400)
                
#                 # Save the Sale object
#                 sale = Sales(
#                     no=no,
#                     bill_number=row["Bill Number"],
#                     date=date,
#                     psr=row["PSR"],
#                     customer_code=row["Customer Code"],
#                     name= customer.description,
#                     city= customer.city,
#                     area= customer.area,
#                     color_making_saler=row["Color Making Saler"],
#                     group= product.group,
#                     product_code=row["Good Code"],
#                     product_name=product.description_ir,
#                     unit=product.unit,
#                     unit2=product.unit_secondary,
#                     original_value=row["The Original Value"],
#                     kg = row['KG'],
#                     original_output_value=row["Original Output Value"],
#                     secondary_output_value=row["Secondary Output Value"],
#                     price=row["Price"],
#                     original_price=row["Original Price"],
#                     discount_percentage=row["Discount Percantage (%)"],
#                     amount_sale=row["Amount Sale"],
#                     discount=row["Discount"],
#                     additional_sales=row["Additional Sales"],
#                     net_sales=row["Net Sales"],
#                     discount_percentage_2=row["Discount Percantage 2(%)"],
#                     real_discount_percentage=row["Real Discount Percantage (%)"],
#                     payment_cash=row["Payment Cash"],
#                     payment_check=row["Payment Check"],
#                     balance=row["Balance"],
#                     saler=row["Saler"],
#                     currency_sepidar=row["Currency-Sepidar"],
#                     dollar_sepidar=row["Dollar-Sepidar"],
#                     currency=row["Currency"],
#                     dollar=row["Dollar"],
#                     manager_rating=row["Manager Rating"],
#                     senior_saler=row["Senior Saler"],
#                     tot_monthly_sales=row["Tot Monthly Sales"],
#                     receipment=row["Receipment"],
#                     ct=row["CT"],
#                     payment_type=row["Payment Type"],
#                     customer_size=row["Customer Size"],
#                     saler_factor=row["Saler Factor"],
#                     prim_percentage=row["Prim Percantage"],
#                     bonus_factor=row["Bonus Factor"],
#                     bonus=row["Bonus"]
#                 )
#                 sale.save()

#                 # Update stock in warehouse
#                 try:
#                     warehouse_item = Warehouse.objects.get(product_code=sale.product_code)
#                     warehouse_item.stock -= sale.original_output_value
#                     warehouse_item.save()
#                 except Warehouse.DoesNotExist:
#                     return JsonResponse({'error': f"No product found with code '{row['Good Code']}' in warehouse"}, status=400)

#                 count += 1
#             return JsonResponse({'message': f"{count} sales data added successfully"}, status=200)

#         except OperationalError as e:
#             return JsonResponse({'error': f"Database error: {str(e)}"}, status=500)
#         except Exception as e:
#             traceback.print_exc()
#             return JsonResponse({'error': str(e)}, status=500)


# class ViewSalesView(APIView):
#     permission_classes = [IsAuthenticated,]
#     authentication_classes = [JWTAuthentication,]
#     def get(self, request, *args, **kwargs):
#         sales = Sales.objects.values().all()
#         sale_list = [[sale['no'], sale['bill_number'], sale['date'].strftime('%Y-%m-%d'), sale['psr'], sale['customer_code'],
#                         sale['name'], sale['city'], sale['area'], sale['color_making_saler'], sale['group'], sale['product_code'], 
#                         sale['product_name'], sale['unit'], sale['unit2'], sale['kg'], sale['original_value'], sale['original_output_value'], 
#                         sale['secondary_output_value'], sale['price'], sale['original_price'], sale['discount_percentage'], sale['amount_sale'],
#                         sale['discount'], sale['additional_sales'], sale['net_sales'], sale['discount_percentage_2'], sale['real_discount_percentage'],
#                         sale['payment_cash'], sale['payment_check'], sale['balance'], sale['saler'], sale['currency_sepidar'], sale['dollar_sepidar'], 
#                         sale['currency'], sale['dollar'], sale['manager_rating'], sale['senior_saler'], sale['tot_monthly_sales'], sale['receipment'], 
#                         sale['ct'], sale['payment_type'], sale['customer_size'], sale['saler_factor'], sale['prim_percentage'], sale['bonus_factor'], 
#                         sale['bonus']]
#                      for sale in sales]
#         return JsonResponse(sale_list, safe=False)

# class DeleteSaleView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def post(self, request, *args, **kwargs):
#         try:
#             no = request.POST.get('no', None)
#             product_code = request.POST.get('product_code', None)
#             original_output_value = request.POST.get('original_output_value', None)
#             Sales.objects.filter(no=no).delete()
#             try:
#                 warehouse_item = Warehouse.objects.get(product_code=product_code)
#                 warehouse_item.stock += float(original_output_value)
#                 warehouse_item.save()
#             except Warehouse.DoesNotExist:
#                 return JsonResponse({'error': f"No product found with code '{product_code}' in warehouse"}, status=400)
#         except Exception as e:
#             return JsonResponse({"error": str(e)}, status=500)
#         return JsonResponse({'message': "Sale object has been successfully deleted"}, status=200)

# class EditSaleView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             data = json.loads(request.body)

#             # Check for required fields
#             for field in ['new_product_code', 'new_customer_code', 'new_original_output_value', 'new_net_sales', 'new_saler', 'new_psr', 'new_date']:
#                 if not data.get(field):
#                     return JsonResponse({'error': f"{field} cannot be empty"}, status=400)

#             # Check for integer fields
#             for field in ['new_product_code', 'new_customer_code']:
#                 try:
#                     if not isinstance(int(data.get(field)), int):
#                         return JsonResponse({'error': f"{field} should be integer"}, status=400)
#                 except Exception as e:
#                     return JsonResponse({'error': f"{field} should be integer"}, status=400)

#             # Check for valid date format
#             try:
#                 new_date = data.get('new_date').split("-")
#                 date = jdatetime.date(int(new_date[0]), int(new_date[1]), int(new_date[2]))
#             except ValueError:
#                 return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
#             except IndexError as e:
#                 return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
#             except Exception as e:
#                 return JsonResponse({'error': str(e)}, status=400)
#             # Check for valid psr value
#             if data.get('new_psr') not in ['P', 'S', 'R']:
#                 return JsonResponse({'error': "Invalid P-S-R value. Allowed values are 'P', 'S', and 'R'."}, status=400)

#             # Check for existing good
#             if not Warehouse.objects.filter(product_code=data.get('new_product_code')).exists():
#                 return JsonResponse({'error': f"No product found with code '{data.get('new_product_code')}' in Warehouse. Please check product code. If there is a new product please add firstly to Warehouse."}, status=400)
#             if not Products.objects.filter(product_code_ir=data.get('new_product_code')).exists():
#                 return JsonResponse({'error': f"No product found with code '{data.get('new_product_code')}'. Please check product code. If there is a new product please add firstly to Products."}, status=400)
#             # Check for existing customer
#             if not Customers.objects.filter(customer_code=data.get('new_customer_code')).exists():
#                 return JsonResponse({'error': f"No customer found with code '{data.get('new_customer_code')}'. Please check customer code. If there is a new customer please add firstly to Customers."}, status=400)

#             # Check for existing saler
#             if not Salers.objects.filter(name=data.get('new_saler')).exists():
#                 return JsonResponse({'error': f"No saler found with name '{data.get('new_saler')}'. Please check saler name. If there is a new saler please add firstly to Salers."}, status=400)
            
#             # Update Sale object
#             old_no = data.get('old_no')
#             if data.get('new_no') and data.get('new_no') != old_no:
#                 if Sales.objects.filter(no=data.get('new_no')).exists():
#                     error_message = f"The sale no '{data.get('new_no')}' already exists in the database."
#                     return JsonResponse({'error': error_message}, status=400)
#             sale = Sales.objects.get(no=old_no)
#             sale.no = data.get('new_no')
#             sale.bill_number = data.get('new_bill_number')
#             sale.date = date
#             sale.psr = data.get('new_psr')
#             sale.customer_code = data.get('new_customer_code')
#             sale.name = data.get('new_name')
#             sale.area = data.get('new_area')
#             sale.city = data.get('new_city')
#             sale.color_making_saler = data.get('new_color_making_saler')
#             sale.group = data.get('new_group')
#             sale.product_code = data.get('new_product_code')
#             sale.product_name = data.get('new_product_name')
#             sale.unit = data.get('new_unit')
#             sale.unit2 = data.get('new_unit2')
#             sale.kg = data.get('new_kg')
#             sale.original_value = data.get('new_original_value')
#             sale.original_output_value = data.get('new_original_output_value')
#             sale.secondary_output_value = data.get('new_secondary_output_value')
#             sale.price = data.get('new_price')
#             sale.original_price = data.get('new_original_price')
#             sale.discount_percentage = data.get('new_discount_percentage')
#             sale.amount_sale = data.get('new_amount_sale')
#             sale.discount = data.get('new_discount')
#             sale.additional_sales = data.get('new_additional_sales')
#             sale.net_sales = data.get('new_net_sales')
#             sale.discount_percentage_2 = data.get('new_discount_percentage_2')
#             sale.real_discount_percentage = data.get('new_real_discount_percentage')
#             sale.payment_cash = data.get('new_payment_cash')
#             sale.payment_check = data.get('new_payment_check')
#             sale.balance = data.get('new_balance')
#             sale.saler = data.get('new_saler')
#             sale.currency_sepidar = data.get('new_currency_sepidar')
#             sale.dollar_sepidar = data.get('new_dollar_sepidar')
#             sale.currency = data.get('new_currency')
#             sale.dollar = data.get('new_dollar')
#             sale.manager_rating = data.get('new_manager_rating')
#             sale.senior_saler = data.get('new_senior_saler')
#             sale.tot_monthly_sales = data.get('new_tot_monthly_sales')
#             sale.receipment = data.get('new_receipment')
#             sale.ct = data.get('new_ct')
#             sale.payment_type = data.get('new_payment_type')
#             sale.customer_size = data.get('new_customer_size')
#             sale.saler_factor = data.get('new_saler_factor')
#             sale.prim_percentage = data.get('new_prim_percentage')
#             sale.bonus_factor = data.get('new_bonus_factor')
#             sale.bonus = data.get('new_bonus')

#             sale.save()

#             return JsonResponse({'message': "Your changes have been successfully saved"}, status=200)

#         except Sales.DoesNotExist:
#             return JsonResponse({'error': "Sale not found!"}, status=400)

#         except ValueError as e:
#             return JsonResponse({'error': str(e)}, status=400)

#         except Exception as e:
#             traceback.print_exc()
#             return JsonResponse({'error': str(e)}, status=500)

# class ExportSalesView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def get(self, request, *args, **kwargs):
#         def set_column_widths(worksheet):
#             for column_cells in worksheet.columns:
#                 length = max(len(str(cell.value)) for cell in column_cells)
#                 worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

#         sales = Sales.objects.all().values()
#         # Create a new workbook and add a worksheet
#         jalali_date= current_jalali_date().strftime('%Y-%m-%d')
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = f"Sales {jalali_date}"
#         # Write the header row
#         header = ['No', 'Bill Number', 'Date', 'PSR', 'Customer Code', 'Name', 'City', 'Area', 'Color Making Saler', 'Group',
#                   'Product Code', 'Product Name', 'Unit', 'Unit2', 'Kg', 'Original Value', 'Original Output Value', 'Secondary Output Value',
#                   'Price', 'Original Price', 'Discount Percentage', 'Amount Sale', 'Discount', 'Additional Sales', 'Net Sales',
#                   'Discount Percentage 2', 'Real Discount Percentage', 'Payment Cash', 'Payment Check', 'Balance', 'Saler', 'Currency Sepidar',
#                   'Dollar Sepidar', 'Currency', 'Dollar', 'Manager Rating', 'Senior Saler', 'Total Monthly Sales', 'Receipment', 'CT',
#                   'Payment Type', 'Customer Size', 'Saler Factor', 'Prim Percentage', 'Bonus Factor', 'Bonus']
#         for col_num, column_title in enumerate(header, 1):
#             cell = ws.cell(row=1, column=col_num)
#             cell.value = column_title
#             cell.font = openpyxl.styles.Font(bold=True)
#             cell.fill = openpyxl.styles.PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')
#             cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='medium'),
#                                                  bottom=openpyxl.styles.Side(style='medium'),
#                                                  left=openpyxl.styles.Side(style='medium'),
#                                                  right=openpyxl.styles.Side(style='medium'))
#         ## Write the data rows
#         for row_num, sale in enumerate(sales, 2):
#             row = [sale['no'], sale['bill_number'], sale['date'].strftime('%Y-%m-%d'), sale['psr'], sale['customer_code'],
#                    sale['name'], sale['city'], sale['area'], sale['color_making_saler'], sale['group'], sale['product_code'], 
#                    sale['product_name'], sale['unit'], sale['unit2'], sale['kg'], sale['original_value'], sale['original_output_value'], 
#                    sale['secondary_output_value'], sale['price'], sale['original_price'], sale['discount_percentage'], sale['amount_sale'],
#                    sale['discount'], sale['additional_sales'], sale['net_sales'], sale['discount_percentage_2'], sale['real_discount_percentage'],
#                    sale['payment_cash'], sale['payment_check'], sale['balance'], sale['saler'], sale['currency_sepidar'], sale['dollar_sepidar'], 
#                    sale['currency'], sale['dollar'], sale['manager_rating'], sale['senior_saler'], sale['tot_monthly_sales'], sale['receipment'], 
#                    sale['ct'], sale['payment_type'], sale['customer_size'], sale['saler_factor'], sale['prim_percentage'], sale['bonus_factor'], 
#                    sale['bonus']]
#             for col_num, cell_value in enumerate(row, 1):
#                 cell = ws.cell(row=row_num, column=col_num)
#                 cell.value = cell_value
#         # Apply some styling to the Excel file
#         for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#             for cell in row:
#                 cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'),
#                                                      bottom=openpyxl.styles.Side(style='thin'),
#                                                      left=openpyxl.styles.Side(style='thin'),
#                                                      right=openpyxl.styles.Side(style='thin'))
#                 cell.alignment = openpyxl.styles.Alignment(horizontal='center')
#         # Set the column widths
#         set_column_widths(ws)

#         # Apply auto filter
#         ws.auto_filter.ref = f"A1:AT{ws.max_row}"


#         # Set the response headers for an Excel file
#         buffer = BytesIO()
#         wb.save(buffer)
#         buffer.seek(0)
#         content = buffer.read()

#         # Encode the content in base64
#         base64_content = base64.b64encode(content).decode()


#         # Send the content and filename in the JSON response
#         return JsonResponse({'filename': f'sales({jalali_date}).xlsx', 'content': base64_content})




# # endregion

# # region Warehouse

# class AddWarehouseView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
    
#     def post(self, request, *args, **kwargs):
#         try:
#             if 'file' not in request.FILES:
#                 raise ValidationError("No file uploaded")
#             file = request.FILES['file']
#             kind = filetype.guess(file.read())
#             if kind is None or kind.mime not in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
#                 return JsonResponse({'error': "The uploaded file is not a valid Excel file"}, status=400)

#             data = pd.read_excel(file)
#             if data.empty:
#                 return JsonResponse({'error': "The uploaded file is empty"}, status=400)
#             count = 0
#             for i, row in data.iterrows():
#                 try:
#                     product_code = int(row["Product Code"])
#                     try:
#                         product = Products.objects.get(product_code_ir = product_code)
#                     except Products.DoesNotExist:
#                         return JsonResponse({'error': f"Product Code: '{product_code}' not found in the Products Table. If there is no mistake, please add '{product_code}' to the Products Table first. "}, status=400)

#                     if Warehouse.objects.filter(product_code=product_code).exists():
#                         continue
#                     count+=1
#                     warehouse_item = Warehouse(product_code=product_code, title=row["Product Title"], unit=row["Unit"], stock=row["Stock"])
#                     warehouse_item.save()
#                 except KeyError as e:
#                     return JsonResponse({'error': f"Column '{e}' not found in the uploaded file"}, status=400)
#             return JsonResponse({'message': f"{count} Warehouse items added successfully"}, status=200)
#         except OperationalError as e:
#             return JsonResponse({'error': f"Database error: {str(e)}"}, status=500)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

# class ViewWarehouseView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def get(self, request, *args, **kwargs):
#         # if not request.user.is_authenticated:
#         #     return HttpResponse(status=401)
#         warehouse_items = Warehouse.objects.values().all()
#         warehouse_list = [[item['product_code'], item['title'], item['unit'], item['stock']] for item in warehouse_items]
#         return JsonResponse(warehouse_list, safe=False)

# class DeleteWarehouseView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def post(self, request, *args, **kwargs):
#         try:
#             product_code = request.POST.get('product_code')
#             Warehouse.objects.filter(product_code=product_code).delete()
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)
#         return JsonResponse({'message': "Warehouse object has been successfully deleted"}, status=200)


# class EditWarehouseView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             data = json.loads(request.body)

#             # Check if old_product_code is provided
#             old_product_code = data.get('old_product_code')
#             if not old_product_code:
#                 return JsonResponse({'error': "Missing required parameter: Product Code"}, status=400)

#             # Get the warehouse item object
#             warehouse_item = Warehouse.objects.get(product_code=old_product_code)

#             # Check if new_product_code value is unique and not empty
#             new_product_code = data.get('new_product_code')
#             if new_product_code and new_product_code != old_product_code:
#                 if not new_product_code:
#                     return JsonResponse({'error': "Product Code cannot be empty!"}, status=400)
#                 if Warehouse.objects.filter(product_code=new_product_code).exists():
#                     return JsonResponse({'error': f"The product code '{new_product_code}' already exists in the warehouse."}, status=400)
#                 else:
#                     try:
#                         product = Products.objects.get(product_code_ir = new_product_code)
#                     except Products.DoesNotExist:
#                         return JsonResponse({'error': f"Your new Product Code: '{new_product_code}' not found in the Products Table. If there is no mistake, please add '{new_product_code}' to the Products Table first. "}, status=400)
#                     warehouse_item.product_code = new_product_code

#             # Update other warehouse item fields
#             for field in ['new_title', 'new_unit', 'new_stock']:
#                 value = data.get(field)
#                 if value is not None and value != '':
#                     updated_field = field[4:]  # Remove the "new_" prefix
#                     setattr(warehouse_item, updated_field, value)
#                 else: 
#                     return JsonResponse({'error': f"{field} cannot be empty!"}, status=400)


#             warehouse_item.save()
#             return JsonResponse({'message': "Your changes have been successfully saved"}, status=200)

#         except Warehouse.DoesNotExist:
#             return JsonResponse({'error': "Warehouse item not found!"}, status=400)

#         except ValueError as e:
#             return JsonResponse({'error': str(e)}, status=400)

#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

# class ExportWarehouseView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def get(self, request, *args, **kwargs):
#         def set_column_widths(worksheet):
#             for column_cells in worksheet.columns:
#                 length = max(len(str(cell.value)) for cell in column_cells)
#                 worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

#         warehouse_items = Warehouse.objects.all().values()
#         # Create a new workbook and add a worksheet
#         jalali_date = current_jalali_date().strftime('%Y-%m-%d')
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = f"Warehouse {jalali_date}"
#         # Write the header row
#         header = ['Product Code', 'Title', 'Unit', 'Stock']
#         for col_num, column_title in enumerate(header, 1):
#             cell = ws.cell(row=1, column=col_num)
#             cell.value = column_title
#             cell.font = openpyxl.styles.Font(bold=True)
#             cell.fill = openpyxl.styles.PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')
#             cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='medium'),
#                                                  bottom=openpyxl.styles.Side(style='medium'),
#                                                  left=openpyxl.styles.Side(style='medium'),
#                                                  right=openpyxl.styles.Side(style='medium'))
#         # Write the data rows
#         for row_num, item in enumerate(warehouse_items, 2):
#             row = [item['product_code'], item['title'], item['unit'], item['stock']]
#             for col_num, cell_value in enumerate(row, 1):
#                 cell = ws.cell(row=row_num, column=col_num)
#                 cell.value = cell_value
#         # Apply some styling to the Excel file
#         for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#             for cell in row:
#                 cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'),
#                                                      bottom=openpyxl.styles.Side(style='thin'),
#                                                      left=openpyxl.styles.Side(style='thin'),
#                                                      right=openpyxl.styles.Side(style='thin'))
#                 cell.alignment = openpyxl.styles.Alignment(horizontal='center')
#         # Set the column widths
#         set_column_widths(ws)

#         # Apply auto filter
#         ws.auto_filter.ref = f"A1:D{ws.max_row}"

#         # Set the response headers for an Excel file
#         buffer = BytesIO()
#         wb.save(buffer)
#         buffer.seek(0)
#         content = buffer.read()

#         # Encode the content in base64
#         base64_content = base64.b64encode(content).decode()

#         # Send the content and filename in the JSON response
#         return JsonResponse({'filename': f'warehouse({jalali_date}).xlsx', 'content': base64_content})



# # endregion

# # region Products

# class AddProductsView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             if 'file' not in request.FILES:
#                 return JsonResponse({'error': "No file uploaded"}, status=400)
#             file = request.FILES['file']
#             kind = filetype.guess(file.read())
#             if kind is None or kind.mime not in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
#                 return JsonResponse({'error': "The uploaded file is not a valid Excel file"}, status=400)

#             data = pd.read_excel(file)
#             if data.empty:
#                 return JsonResponse({'error': "The uploaded file is empty"}, status=400)
#             count = 0
#             for i, row in data.iterrows():
#                 try:
#                     product_code_ir = row["Product Number IR"]
#                     if Products.objects.filter(product_code_ir=product_code_ir).exists():
#                         continue
#                     count+=1
#                     product = Products(
#                         group=row["Group"],
#                         subgroup=row["Subgroup"],
#                         feature=row["Feature"],
#                         product_code_ir=product_code_ir,
#                         product_code_tr=row["Product Number TR"],
#                         description_tr=row["Description TR"],
#                         description_ir=row["Description IR"],
#                         unit=row["Unit"],
#                         unit_secondary=row["Unit Secondary"],
#                         weight = row["Weight"],
#                         currency = row["Currency"],
#                         price= row["Price"]
#                     )
#                     product.save()
#                 except KeyError as e:
#                     return JsonResponse({'error': f"Column '{e}' not found in the uploaded file"}, status=400)
#                 except Exception as e:
#                     return JsonResponse({'error': str(e)}, status=400)
#             return JsonResponse({'message': f"{count} Products data added successfully"}, status=200)
#         except OperationalError as e:
#             return JsonResponse({'error': f"Database error: {str(e)}"}, status=500)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)


# class ViewProductsView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def get(self, request, *args, **kwargs):
#         products = Products.objects.values().all()
#         product_list = [[p['group'], p['subgroup'], p['feature'], p['product_code_ir'], p['product_code_tr'],
#                          p['description_tr'], p['description_ir'], p['unit'], p['unit_secondary'],p['weight'],p['currency'], p['price']] for p in products]
#         return JsonResponse(product_list, safe=False)

# class DeleteProductView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def post(self, request, *args, **kwargs):
#         try:
#             product_code_ir = request.POST.get('product_code_ir')
#             Products.objects.filter(product_code_ir=product_code_ir).delete()
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)
#         return JsonResponse({'message': "Product object has been successfully deleted"}, status=200)

# class EditProductView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             data = json.loads(request.body)

#             old_product_code_ir = data.get('old_product_code_ir')
#             product = Products.objects.get(product_code_ir=old_product_code_ir)

#             # Check if new product_code_ir value is unique
#             new_product_code_ir = data.get('new_product_code_ir')
#             if new_product_code_ir and new_product_code_ir != old_product_code_ir:
#                 if Products.objects.filter(product_code_ir=new_product_code_ir).exists():
#                     return JsonResponse({'error': f"The Product Code IR '{new_product_code_ir}' already exists in the database."}, status=400)
#                 if not new_product_code_ir:
#                     return JsonResponse({'error': "Product Code IR cannot be empty!"}, status=400)
#                 else:
#                     product.product_code_ir = new_product_code_ir

#             # Update other product fields
#             for field in [ 'new_currency', 'new_description_ir', 'new_description_tr', 'new_feature', 'new_group', 'new_price', 'new_product_code_tr', 'new_subgroup', 'new_unit', 'new_unit_secondary', 'new_weight']:
#                 value = data.get(field)
#                 if value is not None and value != '':
#                     updated_field = field[4:]  # Remove the "new_" prefix
#                     setattr(product, updated_field, value)
#                 else:
#                     return JsonResponse({'error': f"The field '{field}' cannot be empty."}, status=400)

#             product.save()
#             return JsonResponse({'message': f"Your changes have been successfully saved."}, status=200)

#         except Products.DoesNotExist:
#             return JsonResponse({'error': "Product not found."}, status=400)

#         except ValueError as e:
#             return JsonResponse({'error': str(e)}, status=400)

#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

# class ExportProductsView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def get(self, request, *args, **kwargs):
#         def set_column_widths(worksheet):
#             for column_cells in worksheet.columns:
#                 length = max(len(str(cell.value)) for cell in column_cells)
#                 worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

#         products = Products.objects.all().values()
#         # Create a new workbook and add a worksheet
#         jalali_date = current_jalali_date().strftime('%Y-%m-%d')
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = f"Products {jalali_date}"
#         # Write the header row
#         header = ['Group', 'Subgroup', 'Feature', 'Product Code (IR)', 'Product Code (TR)', 'Description (TR)', 
#                   'Description (IR)', 'Unit', 'Secondary Unit', 'Weight', 'Currency', 'Price']
#         for col_num, column_title in enumerate(header, 1):
#             cell = ws.cell(row=1, column=col_num)
#             cell.value = column_title
#             cell.font = openpyxl.styles.Font(bold=True)
#             cell.fill = openpyxl.styles.PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')
#             cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='medium'),
#                                                  bottom=openpyxl.styles.Side(style='medium'),
#                                                  left=openpyxl.styles.Side(style='medium'),
#                                                  right=openpyxl.styles.Side(style='medium'))
#         # Write the data rows
#         for row_num, product in enumerate(products, 2):
#             row = [product['group'], product['subgroup'], product['feature'], product['product_code_ir'], 
#                    product['product_code_tr'], product['description_tr'], product['description_ir'], product['unit'], 
#                    product['unit_secondary'], product['weight'], product['currency'], product['price']]
#             for col_num, cell_value in enumerate(row, 1):
#                 cell = ws.cell(row=row_num, column=col_num)
#                 cell.value = cell_value
#         # Apply some styling to the Excel file
#         for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#             for cell in row:
#                 cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'),
#                                                      bottom=openpyxl.styles.Side(style='thin'),
#                                                      left=openpyxl.styles.Side(style='thin'),
#                                                      right=openpyxl.styles.Side(style='thin'))
#                 cell.alignment = openpyxl.styles.Alignment(horizontal='center')
#         # Set the column widths
#         set_column_widths(ws)

#         # Apply auto filter
#         ws.auto_filter.ref = f"A1:L{ws.max_row}"

#         # Set the response headers for an Excel file
#         buffer = BytesIO()
#         wb.save(buffer)
#         buffer.seek(0)
#         content = buffer.read()

#         # Encode the content in base64
#         base64_content = base64.b64encode(content).decode()

#         # Send the content and filename in the JSON response
#         return JsonResponse({'filename': f'products({jalali_date}).xlsx', 'content': base64_content})




# # endregion

# # region Charts

# class ChartView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def post(self, request, *args, **kwargs):
#         #start_date = request.POST.get('start_date')
#         #end_date = request.POST.get('end_date')
#         #product_code = request.POST.get('product_code')
#         #start_date = date(1400,4,1)
#         #end_date = date(1400,4,30)
#         data = Sales.objects.filter(group = "Boya").values('date', 'original_output_value')
#         date_list = [obj['date'] for obj in data]
#         output_value_list = [obj['original_output_value'] for obj in data]
#         response_data = {'date_list': date_list, 'output_value_list': output_value_list}

#         return JsonResponse(response_data, safe=False)

# class ItemListView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def get(self, request, *args, **kwargs):
#         product_codes = Products.objects.values_list('product_code_ir', flat=True)
#         return JsonResponse(list(product_codes), safe=False)
    
#     def post(self, request, *args, **kwargs):
#         # Get the product_title from the POST data

#         data = json.loads(request.body)
#         product_code = data.get('product_code')

#         # Filter Sales by the product_title
#         data = Sales.objects.filter(product_code=product_code).values('date', 'original_output_value')
#         product_name = Products.objects.filter(product_code_ir=product_code).values('description_ir')
#         # Get the original_output_value of each sale
#         date_list = [obj['date'] for obj in data]
#         output_value_list = [obj['original_output_value'] for obj in data]
#         product_name = [obj["product_title"] for obj in product_name]

#         response_data = {'product_name':product_name ,'date_list': date_list, 'output_value_list': output_value_list}

#         # Return the list of output_values as a JSON response
#         return JsonResponse(response_data, safe=False)

# # endregion 



# # region Saler
# class AddSalerView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             data = json.loads(request.body)
#             print(data)
#             jalali_date = data.get("job_start_date").split("-")
#             saler_type = data.get("saler_type")
#             try:
#                 jalali_date = jdatetime.date(int(jalali_date[0]), int(jalali_date[1]), int(jalali_date[2]))
#             except ValueError:
#                 return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
#             except IndexError as e:
#                 return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
#             except Exception as e:
#                 return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
#             if the_man_from_future(jalali_date):
#                 return JsonResponse({'error': "HERE'S THE MAN FROM THE FUTURE TO SAVE US ALL!!!! Job Start Date cannot be future time, please check it :) "}, status=400)
            
#             if saler_type == "Active":
#                 bool_active_saler = True
#                 bool_passive_saler = False
#                 experience_rating = calculate_experience_rating(jalali_date)
#             else:
#                 bool_active_saler = False
#                 bool_passive_saler = True
#                 experience_rating = calculate_passive_saler_experience_rating(jalali_date)
                
#             jalali_current_date = current_jalali_date()
           

#             saler = Salers(
#                 name = data.get("name"),
#                 job_start_date = jalali_date,
#                 manager_performance_rating = 1,
#                 experience_rating = experience_rating,
#                 monthly_total_sales_rating = 1, #will be calculated!!!!!!!!!!!!!!!!!!!!!!
#                 receipment_rating = 1, #will be calculated!!!!!!!!!!!!!!!!!!!!!!
#                 is_active =  True,
#                 is_active_saler = bool_active_saler,
#                 is_passive_saler = bool_passive_saler   
#             )
#             saler.save()
#             return JsonResponse({'message': "Saler added successfully"}, status=200)
#         except IndexError as e:
#             return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD' "}, status=400)
#         except ValueError as e:
#             return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD' "}, status=400)
#         except Exception as e:
#             traceback.print_exc()
#             return JsonResponse({'error': str(e)}, status=500)

# class EditSalerView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             data = json.loads(request.body)
#             print(data)

#             old_data = data.get('old_data')
#             new_data = data.get('new_data')

#             # Check if name is provided
#             name = new_data['name']
#             if not name:
#                 traceback.print_exc()
#                 return JsonResponse({'error': "Missing required parameter: 'name'"}, status=400)
            
#             # Get the saler object
#             saler = Salers.objects.get(id=old_data['id'])
#             print(saler)
            
#             if saler.is_active_saler == True and saler.is_active_saler == False:
#                 # Update other saler fields
#                 for field in ['name', 'job_start_date', 'manager_performance_rating', 'is_active']:
#                     if field == "job_start_date":
#                         try:
#                             new_date =new_data['job_start_date'].split("-")
#                             date = jdatetime.date(int(new_date[0]), int(new_date[1]), int(new_date[2]))
#                             saler.experience_rating = calculate_experience_rating(date)
#                             print("active: ",saler.experience_rating)
#                         except ValueError:
#                             return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
#                         except IndexError as e:
#                             return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
#                         except Exception as e:
#                             traceback.print_exc()
#                             return JsonResponse({'error': str(e)}, status=400)
#                     value = new_data[f'{field}']

#                     if value is not None and value != '':
#                         setattr(saler, field, value)
#                     else:
#                         traceback.print_exc() 
#                         return JsonResponse({'error': "One or more data field is empty!"}, status=400)


#                 saler.save()
#                 return JsonResponse({'message': "Your changes have been successfully saved"}, status=200)
#             else:
#                 try:
#                     # Update other saler fields
#                     for field in ['name', 'job_start_date', 'manager_performance_rating', 'is_active']:
#                         if field == "job_start_date":
#                             try:
#                                 new_date =new_data['job_start_date'].split("-")
#                                 date = jdatetime.date(int(new_date[0]), int(new_date[1]), int(new_date[2]))
#                                 print("date:",date)
#                                 saler.experience_rating = calculate_passive_saler_experience_rating(date)
#                                 print("passive: ",saler.experience_rating)
#                             except ValueError:
#                                 return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
#                             except IndexError as e:
#                                 return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
#                             except Exception as e:
#                                 return JsonResponse({'error': str(e)}, status=400)
#                         value = new_data[f'{field}']

#                         if value is not None and value != '':
#                             setattr(saler, field, value)
#                         else:
#                             traceback.print_exc() 
#                             return JsonResponse({'error': "One or more data field is empty!"}, status=400)


#                     saler.save()
#                     return JsonResponse({'message': "Your changes have been successfully saved"}, status=200)
#                 except Exception as e:
#                     traceback.print_exc()
#                     return JsonResponse({'error': str(e)}, status=500)


#         except Salers.DoesNotExist:
#             traceback.print_exc()
#             return JsonResponse({'error': "Saler not found"}, status=400)

#         except ValueError as e:
#             traceback.print_exc()
#             return JsonResponse({'error': str(e)}, status=400)

#         except Exception as e:
#             traceback.print_exc()
#             return JsonResponse({'error': str(e)}, status=500)



# class CollapsedSalerView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
    
#     def get(self, request, *args, **kwargs):
#         active_salers = Salers.objects.filter(is_deleted=False, is_active_saler=True)
#         print(active_salers)
#         active_salers_list = [[saler.id, saler.name, saler.is_active] for saler in active_salers]
#         passive_salers = Salers.objects.filter(is_deleted=False, is_passive_saler=True)
#         passive_salers_list = [[saler.id, saler.name, saler.is_active] for saler in passive_salers]
        
#         return JsonResponse({"active_salers_list": active_salers_list,
#                              "passive_salers_list": passive_salers_list}, safe=False)


    
#  # Everyday experience rating must be automatically updated !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!   
# class SalerCardView(APIView):
#     permission_classes = (IsAuthenticated,) 
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def post(self, request, *args, **kwargs):
#         data = json.loads(request.body)
#         id = data.get('id')
#         saler = Salers.objects.get(id=id, is_deleted = False)
#         jalali_current_date = current_jalali_date()
#         try:
#             saler_monthly_ratings = SalerMonthlySaleRating.objects.get(name=saler.name, month = jalali_current_date.month, year= jalali_current_date.year )
#             monthly_sale_rating = saler_monthly_ratings.sale_rating
#         except SalerMonthlySaleRating.DoesNotExist:
#             monthly_sale_rating = 1
#         if saler.is_active_saler:
#             response_data = {'id': id , 'name': saler.name, 'job_start_date': saler.job_start_date.strftime('%Y-%m-%d'), 'manager_performance_rating': saler.manager_performance_rating,
#                             'experience_rating': saler.experience_rating, 'monthly_total_sales_rating': monthly_sale_rating, 'receipment_rating':saler.receipment_rating,
#                             'is_active': saler.is_active, "active_or_passive": "Active" }
#         else:
#             response_data = {'id': id , 'name': saler.name, 'job_start_date': saler.job_start_date.strftime('%Y-%m-%d'), 'manager_performance_rating': saler.manager_performance_rating,
#                             'experience_rating': saler.experience_rating, 'monthly_total_sales_rating': monthly_sale_rating, 'receipment_rating':saler.receipment_rating,
#                             'is_active': saler.is_active, "active_or_passive": "Passive" }
#         # Return the list of output_values as a JSON response
#         return JsonResponse(response_data, safe=False)






# class SalerTableView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def get(self, request, *args, **kwargs):
#         salers = Salers.objects.filter(is_deleted=False).values()
#         saler_list = [[s['id'], s['name'], s['job_start_date'].strftime('%Y-%m-%d'), s['manager_performance_rating'],
#                        s['experience_rating'], s['monthly_total_sales_rating'], s['receipment_rating'], s['is_active'], s['is_active_saler'] ,s['is_passive_saler']] for s in salers]
        
#         print(saler_list)
#         return JsonResponse(saler_list, safe=False)

# class DeleteSalerView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def post(self, request, *args, **kwargs):
#         try:    
#             data = json.loads(request.body)
#             id = data.get('id')
#             saler = Salers.objects.get(id=id)
#             print("saler: ", saler)
#             saler.is_deleted = True
#             saler.is_active = False
#             saler.save()

#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)
#         return JsonResponse({'message': "Saler object has been successfully deleted"}, status=200)





# @receiver(pre_save, sender=SalerPerformance)
# def update_month_sale_rating(sender, instance, **kwargs):
#     # Calculate the sale rating based on the updated sale value
#     aggregated_sales = SalerPerformance.objects.filter(
#         name=instance.name,
#         year=instance.year,
#         month=instance.month
#     ).aggregate(monthly_sale=Sum('sale'))
#     monthly_sale = float(aggregated_sales['monthly_sale'] or 0)
#     monthly_sale_rating = calculate_sale_rating(monthly_sale / 10000000)
#     saler_rating, created = SalerMonthlySaleRating.objects.get_or_create(
#         name=instance.name, 
#         year=instance.year,
#         month=instance.month
#         )
#     saler_rating.sale_rating = monthly_sale_rating
#     saler_rating.save()
#     saler = Salers.objects.get(name=instance.name)
#     saler.monthly_total_sales_rating = monthly_sale_rating





# # endregion

# # region SalerPerformance

# @receiver(pre_save, sender=Sales)
# def update_saler_performance_with_add_sale(sender, instance, created, **kwargs):
#     # Get or create the SalerPerformance object
#     saler_performance, _ = SalerPerformance.objects.get_or_create(
#         name=instance.saler,
#         year=instance.date.year,
#         month=instance.date.month,
#         day=instance.date.day
#     )

#     if created:
#         # Update the sale value for the SalerPerformance object
#         saler_performance.sale += float(instance.net_sales)
#         saler_performance.bonus += float(instance.bonus)
#     else:
#         # Check which fields have been updated
#         dirty_fields = instance.get_dirty_fields()

#         # If any of the saler, year, month, or day fields are updated, find the previous SalerPerformance instance, subtract the old values, and update the new SalerPerformance instance
#         if any(field in dirty_fields for field in ['saler', 'date']):
#             old_saler = dirty_fields.get('saler', instance.saler)
#             old_date = dirty_fields.get('date', instance.date)
#             old_saler_performance = SalerPerformance.objects.get(
#                 name=old_saler,
#                 year=old_date.year,
#                 month=old_date.month,
#                 day=old_date.day
#             )

#             # Subtract old values from the old SalerPerformance instance
#             old_saler_performance.sale -= dirty_fields.get('net_sales', instance.net_sales)
#             old_saler_performance.bonus -= dirty_fields.get('bonus', instance.bonus)
#             old_saler_performance.save()

#             # Update the new SalerPerformance instance
#             saler_performance.sale += instance.net_sales
#             saler_performance.bonus += instance.bonus

#         # Update the corresponding attributes of the SalerPerformance instance based on the updated fields
#         else:
#             if 'net_sales' in dirty_fields:
#                 saler_performance.sale += float(instance.net_sales) - dirty_fields['net_sales']
#             if 'bonus' in dirty_fields:
#                 saler_performance.bonus += float(instance.bonus) - dirty_fields['bonus']

#     saler_performance.save()


# @receiver(post_delete, sender=Sales)
# def update_saler_performance_with_delete_sale(sender, instance, **kwargs):
#     # Get the corresponding SalerPerformance object for the sale
#     performance, created = SalerPerformance.objects.get_or_create(
#         name=instance.saler, 
#         year=instance.date.year,
#         month=instance.date.month,
#         day=instance.date.day
#         )

#     # Subtract the net sale amount from the sale field
#     performance.sale -= instance.net_sales
#     performance.bonus -= instance.bonus
#     performance.save()


# class SalerPerformanceView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def get(self, request, *args, **kwargs):
#         saler_performances = SalerPerformance.objects.values().all()
#         saler_performance_list = [[performance['name'], performance['year'], performance['month'],
#                                    performance['day'], performance['sale'], performance['bonus']]
#                                   for performance in saler_performances]
#         return JsonResponse(saler_performance_list, safe=False)



# # endregion

# # region SaleSummary

# @receiver(pre_save, sender=Sales)
# def update_sale_summary_with_add_sale(sender, instance, created, **kwargs):
#     find_month = instance.date.month
#     find_year = instance.date.year
#     find_day = instance.date.day
#     sale_summary, _ = SaleSummary.objects.get_or_create(
#         date=jdatetime.date(int(find_year), int(find_month), int(find_day)),
#         year=find_year,
#         month=find_month,
#         day=find_day
#     )

#     if created:
#         # Add the values of all relevant fields to the corresponding attributes of the SaleSummary instance
#         sale_summary.sale += instance.net_sales
#         sale_summary.dollar_sepidar_sale += instance.dollar_sepidar
#         sale_summary.dollar_sale += instance.dollar
#         sale_summary.kg_sale += instance.kg
#         sale_summary.save()
#     else:
#         # Check which fields have been updated
#         dirty_fields = instance.get_dirty_fields()

#         # If the date field is updated, find the previous SaleSummary instance, subtract the old values, and update the new SaleSummary instance
#         if 'date' in dirty_fields:
#             old_date = dirty_fields['date']
#             old_sale_summary = SaleSummary.objects.get(
#                 year=old_date.year,
#                 month=old_date.month,
#                 day=old_date.day
#             )

#             # Subtract old values from the old SaleSummary instance
#             old_sale_summary.sale -= dirty_fields.get('net_sales', instance.net_sales)
#             old_sale_summary.dollar_sepidar_sale -= dirty_fields.get('dollar_sepidar', instance.dollar_sepidar)
#             old_sale_summary.dollar_sale -= dirty_fields.get('dollar', instance.dollar)
#             old_sale_summary.kg_sale -= dirty_fields.get('kg', instance.kg)
#             old_sale_summary.save()

#             # Update the new SaleSummary instance
#             sale_summary.sale += instance.net_sales
#             sale_summary.dollar_sepidar_sale += instance.dollar_sepidar
#             sale_summary.dollar_sale += instance.dollar
#             sale_summary.kg_sale += instance.kg

#         # Update the corresponding attributes of the SaleSummary instance based on the updated fields
#         else:
#             if 'net_sales' in dirty_fields:
#                 sale_summary.sale += float(instance.net_sales) - dirty_fields['net_sales']
#             if 'dollar_sepidar' in dirty_fields:
#                 sale_summary.dollar_sepidar_sale += float(instance.dollar_sepidar) - dirty_fields['dollar_sepidar']
#             if 'dollar' in dirty_fields:
#                 sale_summary.dollar_sale += float(instance.dollar) - dirty_fields['dollar']
#             if 'kg' in dirty_fields:
#                 sale_summary.kg_sale += float(instance.kg) - dirty_fields['kg']

#         sale_summary.save()

# @receiver(post_delete, sender=Sales)
# def update_sale_summary_with_delete_sale(sender, instance, **kwargs):
#     # Get or create the SaleSummary object
#     find_month = instance.date.month
#     find_year = instance.date.year
#     find_day = instance.date.day

#     sale_summary = SaleSummary.objects.get(
#         date=jdatetime.date(int(find_year), int(find_month), int(find_day))
#     )

#     # Update the sale value for the SaleSummary object
#     sale_summary.sale -= instance.net_sales
#     sale_summary.dollar_sepidar_sale -= instance.dollar_sepidar
#     sale_summary.dollar_sale -= instance.dollar
#     sale_summary.kg_sale -= instance.kg
#     sale_summary.save()

# class SalesReportView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def post(self, request, *args, **kwargs):
#         data = json.loads(request.body)

#         report_type = data.get('report_type')
#         start_date =  data.get('start_date').split("-")
#         end_date =  data.get('end_date').split("-")
       

#         if report_type == 'daily':
#             start_date = jdatetime.date(int(start_date[0]), int(start_date[1]), int(start_date[2]))
#             end_date = jdatetime.date(int(end_date[0]), int(end_date[1]), int(end_date[2]))
#             data = SaleSummary.objects.filter(date__range = [start_date, end_date]).values('date').annotate(total_sales=Sum('sale')).order_by('date')
#             sales_report_list = [[d['date'].strftime('%Y-%m-%d'), d['total_sales']] for d in data]

#         elif report_type == 'monthly':
#             start_year, start_month = int(start_date[0]), int(start_date[1])
#             end_year, end_month = int(end_date[0]), int(end_date[1])
#             sales_report_list = []
            
#             while start_year < end_year or (start_year == end_year and start_month < end_month):
#                 # Get the data for the current month
#                 data = SaleSummary.objects.filter(year=start_year, month=start_month).values('year', 'month').annotate(total_sales=Sum('sale'))

#                 # Format the data for the current month and add it to the sales report list
#                 sales_report_list.extend([[f"{jdatetime.date(start_year, start_month, 1).strftime('%Y-%m-%d')}", d['total_sales']] for d in data])

#                 # Increment the month and year
#                 start_month += 1
#                 if start_month > 12:
#                     start_month = 1
#                     start_year += 1

#             # Sort the sales report list by date
#             sales_report_list.sort()
#         elif report_type == 'yearly':
#             start_year  = int(start_date[0])
#             end_year  = int(end_date[0])
#             sales_report_list = []
            
#             while start_year < end_year :
#                 # Get the data for the current month
#                 data = SaleSummary.objects.filter(year=start_year).values('year').annotate(total_sales=Sum('sale'))

#                 # Format the data for the current month and add it to the sales report list
#                 sales_report_list.extend([[f"{jdatetime.date(start_year, 1, 1).strftime('%Y-%m-%d')}", d['total_sales']] for d in data])

#                 # Increment the month and year
#                 start_year += 1

#             # Sort the sales report list by date
#             sales_report_list.sort()

#         else:
#             data = []
#         return JsonResponse(sales_report_list, safe=False)




# # endregion

# # region MonthlyProductSales

# @receiver(pre_save, sender=Sales)
# def update_monthly_product_sales_with_add_sale(sender, instance, created, **kwargs):
#     # Get or create the MonthlyProductSales object
#     monthly_sale, _ = MonthlyProductSales.objects.get_or_create(
#         product_code=instance.product_code,
#         year=instance.date.year,
#         month=instance.date.month
#     )
#     monthly_sale.date = instance.date
#     monthly_sale.product_name = instance.product_name

#     if created:
#         # Update the MonthlyProductSales instance with the new sales information
#         monthly_sale.piece += instance.original_output_value
#         monthly_sale.sale += instance.net_sales
#     else:
#         # Check which fields have been updated
#         dirty_fields = instance.get_dirty_fields()

#         # If any of the product_code, year, or month fields are updated, find the previous MonthlyProductSales instance, subtract the old values, and update the new MonthlyProductSales instance
#         if any(field in dirty_fields for field in ['product_code', 'date']):
#             old_product_code = dirty_fields.get('product_code', instance.product_code)
#             old_date = dirty_fields.get('date', instance.date)
#             old_monthly_sale = MonthlyProductSales.objects.get(
#                 product_code=old_product_code,
#                 year=old_date.year,
#                 month=old_date.month
#             )

#             # Subtract old values from the old MonthlyProductSales instance
#             old_monthly_sale.piece -= dirty_fields.get('original_output_value', instance.original_output_value)
#             old_monthly_sale.sale -= dirty_fields.get('net_sales', instance.net_sales)
#             old_monthly_sale.save()

#             # Update the new MonthlyProductSales instance
#             monthly_sale.piece += instance.original_output_value
#             monthly_sale.sale += instance.net_sales

#         # Update the corresponding attributes of the MonthlyProductSales instance based on the updated fields
#         else:
#             if 'original_output_value' in dirty_fields:
#                 monthly_sale.piece += float(instance.original_output_value) - dirty_fields['original_output_value']
#             if 'net_sales' in dirty_fields:
#                 monthly_sale.sale += float(instance.net_sales) - dirty_fields['net_sales']

#     monthly_sale.save()



# @receiver(post_delete, sender=Sales)
# def update_monthly_product_sales_with_delete_sale(sender, instance, **kwargs):

#     monthly_sale = MonthlyProductSales.objects.get(
#         product_code=instance.product_code,
#         year=instance.date.year,
#         month= instance.date.month
#         )

#     monthly_sale.product_name = instance.product_name,
#     monthly_sale.piece -= instance.original_output_value
#     monthly_sale.sale -= instance.net_sales
#     monthly_sale.save()

# # endregion


# # region Dashboard #!DASHBOARD PAGE START

# # region Customer Performance

# @receiver(pre_save, sender=Sales)
# def update_customer_performance_with_add_sale(sender, instance, created, **kwargs):
#     # Get or create the CustomerPerformance object
#     find_month = instance.date.month
#     find_year = instance.date.year
#     customer_performance, _ = CustomerPerformance.objects.get_or_create(
#          year=find_year, month=find_month, customer_code=instance.customer_code
#     )

#     # Update the corresponding attributes of the CustomerPerformance instance based on the updated fields
#     customer_performance.customer_name = instance.name
#     customer_performance.customer_area = instance.area

#     if created:
#         # Update the sale value for the CustomerPerformance object
#         customer_performance.sale += instance.net_sales
#         customer_performance.sale_amount += instance.original_output_value
#         customer_performance.dollar += instance.dollar
#         customer_performance.dollar_sepidar += instance.dollar_sepidar
#     else:
#         # Check which fields have been updated
#         dirty_fields = instance.get_dirty_fields()

#         # If any of the year, month, or customer_code fields are updated, find the previous CustomerPerformance instance, subtract the old values, and update the new CustomerPerformance instance
#         if any(field in dirty_fields for field in ['date', 'customer_code']):
#             old_date = dirty_fields.get('date', instance.date)
#             old_customer_code = dirty_fields.get('customer_code', instance.customer_code)
#             old_customer_performance = CustomerPerformance.objects.get(
#                 year=old_date.year, month=old_date.month, customer_code=old_customer_code
#             )

#             # Subtract old values from the old CustomerPerformance instance
#             old_customer_performance.sale -= dirty_fields.get('net_sales', instance.net_sales)
#             old_customer_performance.sale_amount -= dirty_fields.get('original_output_value', instance.original_output_value)
#             old_customer_performance.dollar -= dirty_fields.get('dollar', instance.dollar)
#             old_customer_performance.dollar_sepidar -= dirty_fields.get('dollar_sepidar', instance.dollar_sepidar)
#             old_customer_performance.save()

#             # Update the new CustomerPerformance instance
#             customer_performance.sale += instance.net_sales
#             customer_performance.sale_amount += instance.original_output_value
#             customer_performance.dollar += instance.dollar
#             customer_performance.dollar_sepidar += instance.dollar_sepidar

#         # Update the corresponding attributes of the CustomerPerformance instance based on the updated fields
#         else:
#             if 'net_sales' in dirty_fields:
#                 customer_performance.sale += float(instance.net_sales) - dirty_fields['net_sales']
#             if 'original_output_value' in dirty_fields:
#                 customer_performance.sale_amount += float(instance.original_output_value) - dirty_fields['original_output_value']
#             if 'dollar' in dirty_fields:
#                 customer_performance.dollar += float(instance.dollar) - dirty_fields['dollar']
#             if 'dollar_sepidar' in dirty_fields:
#                 customer_performance.dollar_sepidar += float(instance.dollar_sepidar) - dirty_fields['dollar_sepidar']

#     customer_performance.save()



# @receiver(post_delete, sender=Sales)
# def update_customer_performance_with_delete_sale(sender, instance, **kwargs):
#     # Get or create the CustomerPerformance object
#     find_month = instance.date.month
#     find_year = instance.date.year
#     customer_performance = CustomerPerformance.objects.get(
#             year= find_year, month = find_month, customer_code = instance.customer_code
#     )

#     # Update the sale value for the CustomerPerformance object
#     customer_performance.customer_name = instance.name
#     customer_performance.customer_area = instance.area
#     customer_performance.sale -= instance.net_sales
#     customer_performance.sale_amount -= instance.original_output_value
#     customer_performance.dollar -= instance.dollar
#     customer_performance.dollar_sepidar -= instance.dollar_sepidar
#     customer_performance.save()

# class TopCustomersView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def post(self, request, *args, **kwargs):
#         data = json.loads(request.body)

#         report_type = data.get('report_type')
#         if report_type == 'monthly':
#             top_customers_list = []
            
#             # Get the current month and year using jdatetime library
#             date_month= current_jalali_date().month
#             date_year= current_jalali_date().year
            
#             # Get the data for the current month
#             top_5_customer_data = CustomerPerformance.objects.filter(month=date_month, year=date_year).order_by('-sale')[:5]
            
#             # Calculate the total sales for the current month
#             total_sales = CustomerPerformance.objects.filter(month=date_month, year=date_year).aggregate(total_sales=Sum('sale'))['total_sales']
#             if total_sales is not None:
#                 # Calculate the sales data for the top 5 customers and others
#                 top_customers_list = [[d.customer_name, d.sale] for d in top_5_customer_data]
#                 top_customers_sale_sum = [d[1] for d in top_customers_list]
#                 top_5_customer_total_sale = sum(top_customers_sale_sum)
#                 others_sales = total_sales - top_5_customer_total_sale

#                 # Create a list of sales data for the top 5 customers and others for the pie chart
#                 top_customers_pie_chart = [[d[0], (d[1]/total_sales)*100] for d in top_customers_list]
#                 top_customers_pie_chart.append(["others", (others_sales/total_sales*100)])
#             else:
#                 # Handle the case when there is no sales data available
#                 top_customers_pie_chart = [["No data available", 100]]
        
#         elif report_type == 'yearly':
#             top_customers_list = []
            
#             # Get the current year using jdatetime library
#             date= current_jalali_date().year
            
#             # Get the data for the current year
#             top_5_customer_data = CustomerPerformance.objects.filter(year=date).order_by('-sale')[:5]
            
#             # Calculate the total sales for the current year
#             total_sales = CustomerPerformance.objects.filter(year=date).aggregate(total_sales=Sum('sale'))['total_sales']
#             if total_sales is not None:
#                 # Calculate the sales data for the top 5 customers and others
#                 top_customers_list = [[d.customer_name, d.sale] for d in top_5_customer_data]
#                 top_customers_sale_sum = [d[1] for d in top_customers_list ]
#                 top_5_customer_total_sale = sum(top_customers_sale_sum)
#                 others_sales = total_sales - top_5_customer_total_sale
                
#                 # Create a list of sales data for the top 5 customers and others for the pie chart
#                 top_customers_pie_chart = [[d[0], (d[1]/total_sales)*100] for d in top_customers_list]
#                 top_customers_pie_chart.append(["others", (others_sales/total_sales*100)])
#             else:
#                 # Handle the case when there is no sales data available
#                 top_customers_pie_chart = [["No data available", 100]]
        
#         return JsonResponse({"top_customers_list": top_customers_list,"top_customers_pie_chart": top_customers_pie_chart}, safe=False)

# # endregion


# # region Product Performance

# @receiver(pre_save, sender=Sales)
# def update_product_performance_with_add_sale(sender, instance, created, **kwargs):
#     # Get or create the ProductPerformance object
#     find_month = instance.date.month
#     find_year = instance.date.year
#     product_performance, _ = ProductPerformance.objects.get_or_create(
#          year=find_year, month=find_month, product_code=instance.product_code
#     )

#     # Update the corresponding attributes of the ProductPerformance instance based on the updated fields
#     product_performance.product_name = instance.product_name

#     if created:
#         # Update the sale value for the ProductPerformance object
#         product_performance.sale_amount += instance.original_output_value
#         product_performance.sale += instance.net_sales
#     else:
#         # Check which fields have been updated
#         dirty_fields = instance.get_dirty_fields()

#         # If any of the year, month, or product_code fields are updated, find the previous ProductPerformance instance, subtract the old values, and update the new ProductPerformance instance
#         if any(field in dirty_fields for field in ['date', 'product_code']):
#             old_date = dirty_fields.get('date', instance.date)
#             old_product_code = dirty_fields.get('product_code', instance.product_code)
#             old_product_performance = ProductPerformance.objects.get(
#                 year=old_date.year, month=old_date.month, product_code=old_product_code
#             )

#             # Subtract old values from the old ProductPerformance instance
#             old_product_performance.sale_amount -= dirty_fields.get('original_output_value', instance.original_output_value)
#             old_product_performance.sale -= dirty_fields.get('net_sales', instance.net_sales)
#             old_product_performance.save()

#             # Update the new ProductPerformance instance
#             product_performance.sale_amount += instance.original_output_value
#             product_performance.sale += instance.net_sales

#         # Update the corresponding attributes of the ProductPerformance instance based on the updated fields
#         else:
#             if 'original_output_value' in dirty_fields:
#                 product_performance.sale_amount += float(instance.original_output_value) - dirty_fields['original_output_value']
#             if 'net_sales' in dirty_fields:
#                 product_performance.sale += float(instance.net_sales) - dirty_fields['net_sales']

#     product_performance.save()



# @receiver(post_delete, sender=Sales)
# def update_product_performance_with_delete_sale(sender, instance, **kwargs):
#     # Get or create the ProductPerformance object
#     find_month = instance.date.month
#     find_year = instance.date.year
#     product_performance = ProductPerformance.objects.get(
#             year= find_year, month = find_month, product_code = instance.product_code
#     )

#     # Update the sale value for the ProductPerformance object
#     product_performance.product_name = instance.product_name
#     product_performance.sale_amount -= instance.original_output_value
#     product_performance.sale -= instance.net_sales
#     product_performance.save()

# class TopProductsView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def post(self, request, *args, **kwargs):
#         data = json.loads(request.body)

#         report_type = data.get('report_type')
#         if report_type == 'monthly':
#             top_products_list = []
            
#             # Get the current month and year using jdatetime library
#             date_month= current_jalali_date().month
#             date_year= current_jalali_date().year
            
#             # Get the data for the current month
#             top_5_product_data = ProductPerformance.objects.filter(month=date_month, year=date_year).order_by('-sale')[:5]
            
#             # Calculate the total sales for the current month
#             total_sales = ProductPerformance.objects.filter(month=date_month, year=date_year).aggregate(total_sales=Sum('sale'))['total_sales']
#             if total_sales is not None:
#                 # Calculate the sales data for the top 5 products and others
#                 top_products_list = [[d.product_name, d.sale] for d in top_5_product_data]
#                 top_products_sale_sum = [d[1] for d in top_products_list ]
#                 top_5_product_total_sale = sum(top_products_sale_sum)
#                 others_sales = total_sales - top_5_product_total_sale

#                 # Create a list of sales data for the top 5 products and others for the pie chart
#                 top_products_pie_chart = [[d[0], (d[1]/total_sales)*100] for d in top_products_list]
#                 top_products_pie_chart.append(["others", (others_sales/total_sales*100)])
#             else:
#                 # Handle the case when there is no sales data available
#                 top_products_pie_chart = [["No data available", 100]]
        
#         elif report_type == 'yearly':
#             top_products_list = []
            
#             # Get the current year using jdatetime library
#             date= current_jalali_date().year
            
#             # Get the data for the current year
#             top_5_product_data = ProductPerformance.objects.filter(year=date).order_by('-sale')[:5]
            
#             # Calculate the total sales for the current year
#             total_sales = ProductPerformance.objects.filter(year=date).aggregate(total_sales=Sum('sale'))['total_sales']
#             if total_sales is not None:
#                 # Calculate the sales data for the top 5 products and others
#                 top_products_list = [[d.product_name, d.sale] for d in top_5_product_data]
#                 top_products_sale_sum = [d[1] for d in top_products_list ]
#                 top_5_product_total_sale = sum(top_products_sale_sum)
#                 others_sales = total_sales - top_5_product_total_sale

#                 # Create a list of sales data for the top 5 products and others for the pie chart
#                 top_products_pie_chart = [[d[0], (d[1]/total_sales)*100] for d in top_products_list]
#                 top_products_pie_chart.append(["others", (others_sales/total_sales*100)])
#             else:
#                 # Handle the case when there is no sales data available
#                 top_products_pie_chart = [["No data available", 100]]
        
#         return JsonResponse({"top_products_list": top_products_list,"top_products_pie_chart": top_products_pie_chart}, safe=False)

# # endregion

# # region Current Exchange Rate

# class ExchangeRateAPIView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def get(self, request):
#         try:
#             exchange_rate = get_exchange_rate()
#             response_data = exchange_rate
#         except Exception as e:
#             response_data = {
#                 "error": "There is an error at Current IRR Exchange Rate. Please contact developer to solve it",
#             }

#         return JsonResponse(response_data,safe=False )

# # endregion

# # region Daily Report

# class SalerDataView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def get(self, request, *args, **kwargs):
#         jalali_date_now = current_jalali_date()
#         jalali_date_now_str = jalali_date_now.strftime('%Y-%m-%d')

#         daily_sales = SalerPerformance.objects.filter(
#             year=jalali_date_now.year,
#             month=jalali_date_now.month,
#             day=jalali_date_now.day
#         ).values('name', 'sale')

#         monthly_sales = SalerPerformance.objects.filter(
#             year=jalali_date_now.year,
#             month=jalali_date_now.month
#         ).values('name').annotate(monthly_sale=Sum('sale'))

#         yearly_sales = SalerPerformance.objects.filter(
#             year=jalali_date_now.year
#         ).values('name').annotate(yearly_sale=Sum('sale'))

#         # Combine the data into a single list
#         combined_data = []
#         if not daily_sales:
#             # If no sales for the current day, append a dictionary with name and daily sale values set to zero
#             for monthly_sale in monthly_sales:
#                 name = monthly_sale['name']
#                 yearly_sale = next((item['yearly_sale'] for item in yearly_sales if item['name'] == name), 0)
#                 try:
#                     saler = Salers.objects.get(name=name)
#                     is_active = saler.is_active
#                 except:
#                     is_active = "Left"
#                 combined_data.append([
#                     name,
#                     is_active,
#                     0,  # Set daily sale value to zero
#                     monthly_sale['monthly_sale'] / 10,
#                     yearly_sale / 10
#                 ])
#         else:
#             for daily_sale in daily_sales:
#                 name = daily_sale['name']
#                 try:
#                     saler = Salers.objects.get(name=name)
#                     is_active = saler.is_active
#                 except:
#                     is_active = "Left"
#                 monthly_sale = next((item['monthly_sale'] for item in monthly_sales if item['name'] == name), 0)
#                 yearly_sale = next((item['yearly_sale'] for item in yearly_sales if item['name'] == name), 0)

#                 combined_data.append([
#                     name,
#                     is_active,
#                     daily_sale['sale'] / 10,
#                     monthly_sale / 10,
#                     yearly_sale / 10
#                 ])

#         response_data = {"jalali_date": jalali_date_now_str, "sales_data": combined_data}

#         return JsonResponse(response_data, safe=False)
        

# class TotalDataView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def get(self, request, *args, **kwargs):
#         jalali_date_now = current_jalali_date()
#         jalali_date_now_str = jalali_date_now.strftime('%Y-%m-%d')
        
#         daily_sales = SaleSummary.objects.filter(
#             year=jalali_date_now.year,
#             month=jalali_date_now.month,
#             day=jalali_date_now.day
#         ).values('sale', 'dollar_sepidar_sale', 'dollar_sale', 'kg_sale')
#         daily_sales_array = list(daily_sales.values_list('sale', 'dollar_sepidar_sale', 'dollar_sale', 'kg_sale')[0]) if daily_sales.exists() else [0, 0, 0, 0]
#         # Divide each value in daily_sales_array by 10
#         for i in range(len(daily_sales_array)):
#             daily_sales_array[i] /= 10

#         monthly_sales = SaleSummary.objects.filter(
#             year=jalali_date_now.year,
#             month=jalali_date_now.month
#         ).annotate(monthly_sale=Sum('sale'), monthly_dollar_sepidar_sale=Sum('dollar_sepidar_sale'), monthly_dollar_sale=Sum('dollar_sale'), monthly_kg_sale=Sum('kg_sale') )
#         monthly_sales_array = list(monthly_sales.values_list('monthly_sale', 'monthly_dollar_sepidar_sale', 'monthly_dollar_sale', 'monthly_kg_sale')[0]) if monthly_sales.exists() else [0, 0, 0, 0]
#         # Divide each value in daily_sales_array by 10
#         for i in range(len(monthly_sales_array)):
#             monthly_sales_array[i] /= 10


#         yearly_sales = SaleSummary.objects.filter(
#             year=jalali_date_now.year
#         ).annotate(yearly_sale=Sum('sale'), yearly_dollar_sepidar_sale=Sum('dollar_sepidar_sale'), yearly_dollar_sale=Sum('dollar_sale'), yearly_kg_sale=Sum('kg_sale') )
#         yearly_sales_array = list(yearly_sales.values_list('yearly_sale', 'yearly_dollar_sepidar_sale', 'yearly_dollar_sale', 'yearly_kg_sale')[0]) if yearly_sales.exists() else [0, 0, 0, 0]
#         # Divide each value in daily_sales_array by 10
#         for i in range(len(yearly_sales_array)):
#             yearly_sales_array[i] /= 10
        

#         response_data = { "jalali_date" : jalali_date_now_str, "daily_sales" : daily_sales_array, "monthly_sales" : monthly_sales_array, "yearly_sales" : yearly_sales_array }

#         # Combine the data into a single list
        
        
#         return JsonResponse(response_data, safe=False)
        

# class TotalDataByMonthlyView(View):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def get(self, request, *args, **kwargs):
#         jalali_date_now = current_jalali_date()
#         jalali_date_now_str = jalali_date_now.strftime('%Y-%m-%d')
        
#         monthly_sales = SaleSummary.objects.filter(
#         year=jalali_date_now.year,
#         month__lte=jalali_date_now.month
#         ).values('month').annotate(monthly_sale=Sum('sale'), monthly_dollar_sepidar_sale=Sum('dollar_sepidar_sale'), monthly_dollar_sale=Sum('dollar_sale'), monthly_kg_sale=Sum('kg_sale'))

#         monthly_sales_dict = {}
#         for monthly_sale_obj in monthly_sales:
#             monthly_sales_dict[monthly_sale_obj['month']] = {
#                 'monthly_sale': monthly_sale_obj['monthly_sale'] / 10,
#                 'monthly_dollar_sepidar_sale': monthly_sale_obj['monthly_dollar_sepidar_sale'],
#                 'monthly_dollar_sale': monthly_sale_obj['monthly_dollar_sale'],
#                 'monthly_kg_sale': monthly_sale_obj['monthly_kg_sale']
#             }

#         monthly_sales_data = []
#         for i in range(1, 13):
#             monthly_sales_obj = monthly_sales_dict.get(i)
#             if monthly_sales_obj:
#                 dollar_sale_per_kg = monthly_sales_obj['monthly_dollar_sale'] / monthly_sales_obj['monthly_kg_sale'] if monthly_sales_obj['monthly_kg_sale'] != 0 else 0
#                 monthly_sales_data.append([
#                     monthly_sales_obj['monthly_sale'],
#                     monthly_sales_obj['monthly_dollar_sepidar_sale'],
#                     monthly_sales_obj['monthly_dollar_sale'],
#                     monthly_sales_obj['monthly_kg_sale'],
#                     dollar_sale_per_kg
#                 ])
#             elif i <= jalali_date_now.month:
#                 monthly_sales_data.append([0, 0, 0, 0, 0])

#         monthly_sales_array = monthly_sales_data if len(monthly_sales_data) == 12 else monthly_sales_data + [[0, 0, 0, 0, 0]] * (12 - len(monthly_sales_data))
#         return JsonResponse(monthly_sales_array, safe=False)


# # endregion

# # region Customer Area 

# class CustomerAreaPieChartView(View):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def post(self, request, *args, **kwargs):
#         data = json.loads(request.body)
        
#         report_type = data.get('report_type')
#         date_month= current_jalali_date().month
#         date_year= current_jalali_date().year
        
#         table_data = []
#         if report_type == 'monthly':
#             data = CustomerPerformance.objects.filter(year=date_year, month=date_month).values('customer_area').annotate(total_dollar=Sum('dollar'))
#         else:  # default is 'yearly'
#             data = CustomerPerformance.objects.filter(year=date_year).values('customer_area').annotate(total_dollar=Sum('dollar'))

        
#         total_dollar = sum([item['total_dollar'] for item in data])
#         if total_dollar is not None and total_dollar != 0:
#             table_data = [[item['customer_area'], item['total_dollar']] for item in data]
#             chart_data_percent = [[item['customer_area'], item['total_dollar'] / total_dollar * 100] for item in data]
            
#         else:
#             # Handle the case when there is no sales data available
#             chart_data_percent = [["No data available", 100]]
        
#         return JsonResponse({"table_data": table_data, "chart_data_percent": chart_data_percent}, safe=False)

# # endregion

# # endregion #!DASHBOARD PAGE END

# # region ROP
# @receiver(pre_save, sender=Warehouse)
# def create_rop_for_warehouse(sender, instance, created, **kwargs):
#     if created:
#         try:
#             product = Products.objects.get(product_code_ir = instance.product_code)
            
#             rop, created = ROP.objects.get_or_create(
#             group = product.group,
#             subgroup = product.subgroup,
#             feature = product.feature,
#             new_or_old_product = 0,
#             related = None,
#             origin = None,
#             product_code_ir = instance.product_code,
#             product_code_tr = product.product_code_tr,
#             dont_order_again = None,
#             description_tr = product.description_tr,
#             description_ir = product.description_ir,
#             unit = product.unit,
#             weight = product.weight,
#             unit_secondary = product.unit_secondary,
#             price = product.price,
#             #color_making_room_1400 = None,
#             avarage_previous_year = None,
#             month_1 = None,
#             month_2 = None,
#             month_3 = None,
#             month_4 = None,
#             month_5 = None,
#             month_6 = None,
#             month_7 = None,
#             month_8 = None,
#             month_9 = None,
#             month_10 = None,
#             month_11 = None,
#             month_12 = None,
#             total_sale = None,
#             warehouse = None,
#             goods_on_the_road = None,
#             total_stock_all = None,
#             total_month_stock = None,
#             standart_deviation = None,
#             lead_time = None,
#             product_coverage_percentage = None,
#             demand_status = None,
#             safety_stock = None,
#             rop = None,
#             monthly_mean = None,
#             new_party = None,
#             cycle_service_level = None,
#             total_stock = None,
#             need_prodcuts = None,
#             over_stock = None,
#             calculated_need = None,
#             calculated_max_stock = None,
#             calculated_min_stock = None,
#             )

#             rop.save()

#         except Exception as e:
#             return JsonResponse({'error': str(e)})

# @receiver(pre_save, sender=Sales)
# def update_rop_for_sales_add_or_edit(sender, instance, created, **kwargs):
#         try:
#             product = Products.objects.get(product_code_ir = instance.product_code)
#             warehouse = Warehouse.objects.get(product_code=instance.product_code)
#             product_performance = ProductPerformance.objects.filter(product_code =instance.product_code )
#             jalali_date_now = current_jalali_date()
#             jalali_date_now_year = int(jalali_date_now.year)
#             jalali_date_now_month = int(jalali_date_now.month)
#             jalali_date_previous_year = jalali_date_now_year-1
#             product_performance_previous_year = product_performance.filter(year=jalali_date_previous_year)
#             # calculate the average sale amount
#             average_sale_amount_previous_year = product_performance_previous_year.aggregate(avg_sale=Avg('sale_amount'))['avg_sale']
            
#             rop = ROP.objects.get(
#                 product_code_ir = instance.product_code
#                 )
            
#             # rop.group = product.group,
#             # rop.subgroup = product.subgroup,
#             # rop.feature = product.feature,
#             # rop.new_or_old_product = 0,
#             # rop.related = None,
#             # rop.origin = None,
#             # rop.product_code_ir = instance.product_code,
#             # rop.product_code_tr = product.product_code_tr,
#             # rop.dont_order_again = None,
#             # rop.description_tr = product.description_tr,
#             # rop.description_ir = product.description_ir,
#             # rop.unit = product.unit,
#             # rop.weight = product.weight,
#             # rop.unit_secondary = product.unit_secondary,
#             # rop.price = product.price,
#             # #rop.color_making_room_1400 = None,
#             rop.avarage_previous_year = average_sale_amount_previous_year
            
#             # amount sales from month_1 to month_12
#             for month_number in range(1, 13):
#                 # Filter the `ProductPerformance` objects by the current year and month.
#                 product_performance_current_month = product_performance.filter(year=jalali_date_now_year, month=month_number)

#                 # Calculate the total sale amount for the current month.
#                 total_sale_current_month = product_performance_current_month.aggregate(total_sale=Coalesce(Sum('sale_amount', output_field=models.FloatField()), float(0)))['total_sale']


#                 # Set the total sale amount for the current month for the corresponding `ROP` month field.
#                 setattr(rop, f'month_{month_number}', total_sale_current_month)
#             rop.total_sale = product_performance.filter(year=jalali_date_now_year).aggregate(total_sale=Coalesce(Sum('sale_amount'), float(0)))['total_sale']
#             rop.warehouse = warehouse.stock
#             rop.goods_on_the_road = float(0) #! goods on road values will be updated!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#             rop.total_stock_all = rop.warehouse+ rop.goods_on_the_road
#             rop.monthly_mean = rop.total_sale/int(jalali_date_now.month)
#             try: 
#                 rop.total_month_stock = rop.total_stock_all/rop.monthly_mean
#             except Exception as e:
#                 rop.total_month_stock = 0
            
#             # Get the values of `month_1` to the current month.
#             values = [getattr(rop, f'month_{i}', 0) for i in range(1, jalali_date_now_month + 1)]
#             # Calculate the standard deviation.
#             if len(values) > 0:
#                 standard_deviation = statistics.stdev(values)
#             else:
#                 standard_deviation = 0

#             # Set the `standart_deviation` field of the `ROP` instance.
#             rop.standart_deviation = standard_deviation
            
#             rop.lead_time = 2 #! will be given by user
#             rop.product_coverage_percentage = 95 #! will be given by user
#             rop.demand_status =  rop.standart_deviation * (rop.lead_time)**0.5

#             #safety stock
#             try:
#                 result = norm.ppf(rop.product_coverage_percentage) * rop.demand_status
#                 result_rounded_up = math.ceil(result)           
#                 rop.safety_stock = result_rounded_up
#             except:
#                 rop.safety_stock = float(0)

#             rop.rop = (rop.lead_time * rop.monthly_mean) + rop.safety_stock 
#             #rop.monthly_mean = rop.total_sale/int(jalali_date_now.month)
#             rop.new_party = rop.lead_time * rop.monthly_mean
            
#             # cycle_service_level
#             try:
#                 # Calculate the PDF of the normal distribution.
#                 pdf_value = norm.pdf(rop.rop, loc=rop.monthly_mean, scale=rop.demand_status)

#                 # Set the result to the PDF.
#                 rop.cycle_service_level = pdf_value
#             except:
#                 # If an error occurs, set the result to 0.
#                 rop.cycle_service_level = 0
#             rop.total_stock = rop.total_stock_all
            
#             # need_products
#             if rop.rop >= rop.total_stock:
#                 rop.need_prodcuts = rop.rop + rop.new_party - rop.total_stock
#             else:
#                  rop.need_prodcuts = 0

#             # over stock
#             if rop.total_stock_all > (1.2*(rop.safety_stock + rop.new_party)):  #! Stock Over Factor will be declared and it will produced by user
#                 rop.over_stock = 1
#             else:
#                 rop.over_stock = 0
            
#             rop.calculated_need = rop.need_prodcuts #! previous year is needed??????
            
#             # calculated max stock #! previous year is needed??????
#             try:
#                 rop.calculated_max_stock = (rop.rop + rop.new_party)/rop.monthly_mean
#             except:
#                 rop.calculated_max_stock = 0

#             # calculated min stock #! previous year is needed??????
#             try:
#                 rop.calculated_min_stock = rop.rop /rop.monthly_mean
#             except:
#                 rop.calculated_min_stock = 0
#             rop.save()
#         except Products.DoesNotExist:
#             pass
#         except Exception as e:
#             return JsonResponse({'error': str(e)})

# @receiver(post_delete, sender=Sales)
# def update_rop_for_sales_delete(sender, instance, **kwargs):
#         try:
#             product = Products.objects.get(product_code_ir = instance.product_code)
#             warehouse = Warehouse.objects.get(product_code=instance.product_code)
#             product_performance = ProductPerformance.objects.filter(product_code =instance.product_code )
#             jalali_date_now = current_jalali_date()
#             jalali_date_now_year = int(jalali_date_now.year)
#             jalali_date_now_month = int(jalali_date_now.month)
#             jalali_date_previous_year = jalali_date_now_year-1
#             product_performance_previous_year = product_performance.filter(year=jalali_date_previous_year)
#             # calculate the average sale amount
#             average_sale_amount_previous_year = product_performance_previous_year.aggregate(avg_sale=Avg('sale_amount'))['avg_sale']
            
#             rop = ROP.objects.get(
#                 product_code_ir = instance.product_code
#                 )
            
#             # rop.group = product.group,
#             # rop.subgroup = product.subgroup,
#             # rop.feature = product.feature,
#             # rop.new_or_old_product = 0,
#             # rop.related = None,
#             # rop.origin = None,
#             # rop.product_code_ir = instance.product_code,
#             # rop.product_code_tr = product.product_code_tr,
#             # rop.dont_order_again = None,
#             # rop.description_tr = product.description_tr,
#             # rop.description_ir = product.description_ir,
#             # rop.unit = product.unit,
#             # rop.weight = product.weight,
#             # rop.unit_secondary = product.unit_secondary,
#             # rop.price = product.price,
#             # #rop.color_making_room_1400 = None,
#             rop.avarage_previous_year = average_sale_amount_previous_year
            
#             # amount sales from month_1 to month_12
#             for month_number in range(1, 13):
#                 # Filter the `ProductPerformance` objects by the current year and month.
#                 product_performance_current_month = product_performance.filter(year=jalali_date_now_year, month=month_number)

#                 # Calculate the total sale amount for the current month.
#                 total_sale_current_month = product_performance_current_month.aggregate(total_sale=Coalesce(Sum('sale_amount', output_field=models.FloatField()), float(0)))['total_sale']


#                 # Set the total sale amount for the current month for the corresponding `ROP` month field.
#                 setattr(rop, f'month_{month_number}', total_sale_current_month)
#             rop.total_sale = product_performance.filter(year=jalali_date_now_year).aggregate(total_sale=Coalesce(Sum('sale_amount'), float(0)))['total_sale']
#             rop.warehouse = warehouse.stock
#             rop.goods_on_the_road = float(0) #! goods on road values will be updated!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#             rop.total_stock_all = rop.warehouse+ rop.goods_on_the_road
#             rop.monthly_mean = rop.total_sale/int(jalali_date_now.month)
#             try: 
#                 rop.total_month_stock = rop.total_stock_all/rop.monthly_mean
#             except Exception as e:
#                 rop.total_month_stock = 0
            
#             # Get the values of `month_1` to the current month.
#             values = [getattr(rop, f'month_{i}', 0) for i in range(1, jalali_date_now_month + 1)]
#             # Calculate the standard deviation.
#             if len(values) > 0:
#                 standard_deviation = statistics.stdev(values)
#             else:
#                 standard_deviation = 0

#             # Set the `standart_deviation` field of the `ROP` instance.
#             rop.standart_deviation = standard_deviation
            
#             rop.lead_time = 2 #! will be given by user
#             rop.product_coverage_percentage = 95 #! will be given by user
#             rop.demand_status =  rop.standart_deviation * (rop.lead_time)**0.5

#             #safety stock
#             try:
#                 result = norm.ppf(rop.product_coverage_percentage) * rop.demand_status
#                 result_rounded_up = math.ceil(result)           
#                 rop.safety_stock = result_rounded_up
#             except:
#                 rop.safety_stock = float(0)

#             rop.rop = (rop.lead_time * rop.monthly_mean) + rop.safety_stock 
#             #rop.monthly_mean = rop.total_sale/int(jalali_date_now.month)
#             rop.new_party = rop.lead_time * rop.monthly_mean
            
#             # cycle_service_level
#             try:
#                 # Calculate the PDF of the normal distribution.
#                 pdf_value = norm.pdf(rop.rop, loc=rop.monthly_mean, scale=rop.demand_status)

#                 # Set the result to the PDF.
#                 rop.cycle_service_level = pdf_value
#             except:
#                 # If an error occurs, set the result to 0.
#                 rop.cycle_service_level = 0
#             rop.total_stock = rop.total_stock_all
            
#             # need_products
#             if rop.rop >= rop.total_stock:
#                 rop.need_prodcuts = rop.rop + rop.new_party - rop.total_stock
#             else:
#                  rop.need_prodcuts = 0

#             # over stock
#             if rop.total_stock_all > (1.2*(rop.safety_stock + rop.new_party)):  #! Stock Over Factor will be declared and it will produced by user
#                 rop.over_stock = 1
#             else:
#                 rop.over_stock = 0
            
#             rop.calculated_need = rop.need_prodcuts #! previous year is needed??????
            
#             # calculated max stock #! previous year is needed??????
#             try:
#                 rop.calculated_max_stock = (rop.rop + rop.new_party)/rop.monthly_mean
#             except:
#                 rop.calculated_max_stock = 0

#             # calculated min stock #! previous year is needed??????
#             try:
#                 rop.calculated_min_stock = rop.rop /rop.monthly_mean
#             except:
#                 rop.calculated_min_stock = 0
#             rop.save()
#         except Products.DoesNotExist:
#             pass
#         except Exception as e:
#             return JsonResponse({'error': str(e)})

# class ROPView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def post(self, request, *args, **kwargs):
#         data = json.loads(request.body)
#         print(data)
#         product_code = data.get('product_code')
#         lead_time = int(data.get('lead_time'))
#         service_level = float(data.get('service_level'))
#         forecast_period = int(data.get('forecast_period'))
#         product_values = ProductPerformance.objects.filter(product_code=product_code)
       
#         try:
#             last_sales = MonthlyProductSales.objects.filter(product_code=product_code).values("date").latest("date")
#             last_sale_date = last_sales["date"].strftime('%Y-%m-%d')
#             jalali_date = current_jalali_date()
#             jalali_date_str = jalali_date.strftime('%Y-%m-%d').split("-")
#         except MonthlyProductSales.DoesNotExist:
#             return JsonResponse({"error" : f"There is no product sales data with product code: {product_code} "}, status=400)
#         try: 
#             warehouse = Warehouse.objects.get(product_code = product_code)
#             stock = warehouse.stock
#         except Warehouse.DoesNotExist:
#             print("sadasdaasd")
#             return JsonResponse({"error" : f"There is no product in warehouse with product code: {product_code} "}, status=400)
        
#         dates_for_sales = [jdatetime.date(item.year, item.month, 1).strftime('%Y-%m-%d') for item in product_values]
#         sales = [item.sale_amount for item in product_values]
#         product_values = [[1, item.month, item.year, item.product_code, item.sale_amount] for item in product_values]
        
#         holt_all_sales, holt_prev_sales, holt_future_sales, holt_future_stocks, holt_order_flag, holt_safety_stock, holt_rop, holt_order = get_model("holt", True, jalali_date_str, product_code, product_values, stock, lead_time, service_level, 12, forecast_period )
#         exp_all_sales, exp_prev_sales, exp_future_sales, exp_future_stocks, exp_order_flag, exp_safety_stock, exp_rop, exp_order = get_model("exp", True, jalali_date_str, product_code, product_values, stock, lead_time, service_level, 12, forecast_period )
#         avrg_all_sales, avrg_prev_sales, avrg_future_sales, avrg_future_stocks, avrg_order_flag, avrg_safety_stock, avrg_rop, avrg_order = get_model("average", True, jalali_date_str, product_code, product_values, stock, lead_time, service_level, 12, forecast_period )
#         if avrg_order_flag == False:
#             avrg_order = 0
#         if exp_order_flag == False:
#             exp_order = 0
#         if holt_order_flag == False:
#             holt_order = 0
#         avrg_future_forecast_dates = generate_future_forecast_dates(len(avrg_future_sales))
#         exp_future_forecast_dates = generate_future_forecast_dates(len(exp_future_sales))
#         holt_future_forecast_dates = generate_future_forecast_dates(len(holt_future_sales))
#         print("avrg_future_forecast_dates: ", avrg_future_forecast_dates)
#         print("holt_future_forecast_dates: ", holt_future_forecast_dates)
#         print("holt_order: ", holt_order)
#         item = ROP.objects.get(product_code_ir = product_code)
#         rop_list = rop_list = [
#                 item.group,
#                 item.subgroup,
#                 item.feature,
#                 item.new_or_old_product,
#                 item.related,
#                 item.origin,
#                 item.product_code_ir,
#                 item.product_code_tr,
#                 item.dont_order_again,
#                 item.description_tr,
#                 item.description_ir,
#                 item.unit,
#                 item.weight,
#                 item.unit_secondary,
#                 item.price,
#                 item.avarage_previous_year,
#                 item.month_1,
#                 item.month_2,
#                 item.month_3,
#                 item.month_4,
#                 item.month_5,
#                 item.month_6,
#                 item.month_7,
#                 item.month_8,
#                 item.month_9,
#                 item.month_10,
#                 item.month_11,
#                 item.month_12,
#                 item.total_sale,
#                 item.warehouse,
#                 item.goods_on_the_road,
#                 item.total_stock_all,
#                 item.total_month_stock,
#                 item.standart_deviation,
#                 item.lead_time,
#                 item.product_coverage_percentage,
#                 item.demand_status,
#                 item.safety_stock,
#                 item.rop,
#                 item.monthly_mean,
#                 item.new_party,
#                 item.cycle_service_level,
#                 item.total_stock,
#                 item.need_prodcuts,
#                 item.over_stock,
#                 item.calculated_need,
#                 item.calculated_max_stock,
#                 item.calculated_min_stock,
#             ]
#         return JsonResponse({'rop_list': rop_list, 'stock': stock,
#                             'avrg_dates_for_sales': dates_for_sales,
#                             'avrg_sales': sales,
#                             'avrg_future_forecast_dates': avrg_future_forecast_dates, 
#                             'avrg_future_sales': avrg_future_sales, 
#                             'avrg_order_flag': avrg_order_flag, 
#                             'avrg_order': avrg_order,
#                             'holt_dates_for_sales': dates_for_sales,
#                             'holt_sales': sales,
#                             'holt_future_forecast_dates': holt_future_forecast_dates, 
#                             'holt_future_sales': holt_future_sales.tolist(), 
#                             'holt_order_flag': holt_order_flag, 
#                             'holt_order': holt_order,
#                             'exp_dates_for_sales': dates_for_sales,
#                             'exp_sales': sales,
#                             'exp_future_forecast_dates': exp_future_forecast_dates, 
#                             'exp_future_sales': exp_future_sales.tolist(), 
#                             'exp_order_flag': exp_order_flag, 
#                             'exp_order': exp_order }, safe=False)



# # endregion

# # region OrderList

# @receiver(pre_save, sender=Sales)
# def create_sales_signal(sender, instance, created, **kwargs):
#     if created:
#         models = ['average', 'holt', 'exp']
#         order_flags = {'average': False, 'holt': False, 'exp': False}
#         orders = {'average': 0, 'holt': 0, 'exp': 0}
#         product_code = instance.product_code
#         product_values = ProductPerformance.objects.filter(product_code=product_code)
#         product_values = [[1, item.month, item.year, item.product_code, item.sale_amount] for item in product_values]
#         weight = Products.objects.get(product_code_ir = product_code).weight
#         try:
#             last_sales = MonthlyProductSales.objects.filter(product_code=product_code).values("date").latest("date")
#             last_sale_date = last_sales["date"].strftime('%Y-%m-%d')
#             jalali_date = current_jalali_date()
#             jalali_date_str = jalali_date.strftime('%Y-%m-%d').split("-")
#         except MonthlyProductSales.DoesNotExist:
#             return JsonResponse({"error" : f"There is no product sales data with product code: {product_code} "})
#         try: 
#             warehouse = Warehouse.objects.get(product_code = product_code)
#             stock = warehouse.stock
#         except Warehouse.DoesNotExist:
#             return JsonResponse({"error" : f"There is no product in warehouse with product code: {product_code} "})

#         for model in models:
#             all_sales, prev_sales, future_sales, future_stocks, order_flag, safety_stock, rop, order = get_model(
#                 model, True, jalali_date_str, product_code, product_values, stock, 3, 0.99,
#                 12, 3
#             )

#             if order_flag:
#                 is_active = True
#                 order_flags[model] = True
#                 orders[model] = order

#         is_active = any(order_flags.values())

#         try:
#             existing_order_list = OrderList.objects.get(product_code=product_code, is_active=True)
#         except OrderList.DoesNotExist:
#             existing_order_list = None

#         if is_active:
#             if existing_order_list:
#                 # Update the existing OrderList object
#                 existing_order_list.order_flag_avrg = order_flags['average']
#                 existing_order_list.order_flag_exp = order_flags['exp']
#                 existing_order_list.order_flag_holt = order_flags['holt']
#                 existing_order_list.order_avrg = orders['average']
#                 existing_order_list.order_exp = orders['exp']
#                 existing_order_list.order_holt = orders['holt']
#                 existing_order_list.current_date = jalali_date
#                 existing_order_list.current_stock = stock
#                 existing_order_list.weight = weight
#                 existing_order_list.average_sale = np.mean(all_sales)
#                 existing_order_list.is_ordered = False
#                 existing_order_list.save()
#             else:
#                 # Create a new OrderList object
#                 order_list = OrderList(
#                     product_code=product_code,
#                     order_flag_avrg=order_flags['average'],
#                     order_flag_exp=order_flags['exp'],
#                     order_flag_holt=order_flags['holt'],
#                     order_avrg=orders['average'],
#                     order_exp=orders['exp'],
#                     order_holt=orders['holt'],
#                     current_date=jalali_date,
#                     current_stock=stock,
#                     weight=weight,
#                     average_sale=np.mean(all_sales),
#                     is_active=is_active,
#                     is_ordered=False
#                 )
#                 order_list.save()




# class OrderListView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def get(self, request, *args, **kwargs):
#         order_lists = OrderList.objects.filter(is_active = True).values() #! False True olacak
#         order_list_data = [
#             [o['id'], o['current_date'].strftime('%Y-%m-%d'), o['product_code'], o['weight'], o['average_sale'], o['current_stock'], o['order_avrg'], o['order_exp'], o['order_holt'],
#               o['decided_order']] for o in order_lists
#         ]
#         return JsonResponse(order_list_data, safe=False)

# class EditOrderListView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)
#     def post(self, request, *args, **kwargs):
#         # Expecting the updated data to be sent as a list of dictionaries
#         updated_order_list_data = json.loads(request.body)
        

#         try:
            
#             order = OrderList.objects.get(id=updated_order_list_data['id'])           
#             order.current_date = updated_order_list_data['current_date']
#             order.product_code = updated_order_list_data['product_code']
#             order.weight = updated_order_list_data['weight']
#             order.average_sale = updated_order_list_data['average_sale']
#             order.current_stock = updated_order_list_data['current_stock']
#             order.order_avrg = updated_order_list_data['order_avrg']
#             order.order_exp = updated_order_list_data['order_exp']
#             order.order_holt = updated_order_list_data['order_holt']
#             order.decided_order = float(updated_order_list_data['decided_order'])

#             if order.decided_order > 0:
#                 order.is_active = False
#                 order.is_ordered = True

#             order.save()
#         except OrderList.DoesNotExist:
#             return JsonResponse({"error": "Order not found"}, status=404)

#         return JsonResponse({'message': "Your order successfully send to the Goods On Road."}, status=200)


# # endregion

# # region GoodsOnRoad

# @receiver(pre_save, sender=OrderList)
# def update_goods_on_road(sender, instance, created, **kwargs):
#     if not created and instance.is_ordered:
#         product = Products.objects.get(product_code_ir=instance.product_code)
#         goods_on_road, created = GoodsOnRoad.objects.get_or_create(product_code=instance.product_code, is_terminated=False)
#         goods_on_road.product_name_tr = product.description_tr
#         goods_on_road.product_name_ir = product.description_ir
#         goods_on_road.decided_order = instance.decided_order
#         goods_on_road.weight = instance.weight
#         goods_on_road.truck_id = None
#         goods_on_road.is_ordered = instance.is_ordered
#         goods_on_road.is_on_truck = False
#         goods_on_road.save()

# class GoodsOnRoadView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def get(self, request, *args, **kwargs):
#         goods_on_road = GoodsOnRoad.objects.filter(is_terminated=False).values()
#         goods_on_road_data = [
#             [g['product_code'], g['product_name_tr'], g['product_name_ir'], g['decided_order'], g['weight'], g['truck_name']]
#             for g in goods_on_road
#         ]
#         return JsonResponse(goods_on_road_data, safe=False)

# class ApproveProductsToOrderView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             data = json.loads(request.body)

#             if not data.get("truck_name"):
#                 return JsonResponse({'error': "Truck Name cannot be empty."}, status=400)

#             truck = Trucks.objects.filter(truck_name=data.get("truck_name"), is_ordered=False).first()
#             if not truck:
#                 return JsonResponse({'error': "No active truck found with the given Name."}, status=404)

#             if float(data.get("decided_order", 0)) <= 0:
#                 return JsonResponse({'error': "Decided order cannot be equal to or smaller than zero."}, status=400)

#             goods_on_road = GoodsOnRoad.objects.get(product_code=data['product_code'], is_terminated=False)
#             goods_on_road.truck_name = data['truck_name']
#             goods_on_road.decided_order = data['decided_order']
#             goods_on_road.is_terminated = True
#             goods_on_road.is_on_truck = True
#             goods_on_road.save()

#             return JsonResponse({'message': "GoodsOnRoad object updated successfully."}, status=200)
#         except GoodsOnRoad.DoesNotExist:
#             return JsonResponse({'error': "GoodsOnRoad object not found."}, status=404)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

# # endregion

# # region Trucks

# class AddTruckView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             data = json.loads(request.body)

#             # Check if a truck with the same name and is_arrived=False already exists
#             existing_truck = Trucks.objects.filter(truck_name=data.get("truck_name"), is_arrived=False)
#             if existing_truck.exists():
#                 return JsonResponse({'error': "A truck with the same name already exists and has not arrived yet."}, status=400)

#             estimated_order_date_jalali = data.get("estimated_order_date").split("-")
#             estimated_arrival_date_jalali = data.get("estimated_arrival_date").split("-")

#             try:
#                 estimated_order_date = jdatetime.date(int(estimated_order_date_jalali[0]), int(estimated_order_date_jalali[1]), int(estimated_order_date_jalali[2]))
#                 estimated_arrival_date = jdatetime.date(int(estimated_arrival_date_jalali[0]), int(estimated_arrival_date_jalali[1]), int(estimated_arrival_date_jalali[2]))
#             except Exception as e:
#                 return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)

#             truck = Trucks(
#                 truck_name=data.get("truck_name"),
#                 estimated_order_date=estimated_order_date,
#                 estimated_arrival_date=estimated_arrival_date,
#                 is_arrived=False,
#                 is_ordered= False,
#                 is_waiting = True
#             )
#             truck.save()
#             return JsonResponse({'message': "Truck added successfully"}, status=200)
#         except IndexError:
#             return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD' "}, status=400)
#         except ValueError:
#             return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD' "}, status=400)
#         except Exception as e:
#              return JsonResponse({'error': str(e)}, status=500)

# class WaitingTrucksView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def get(self, request, *args, **kwargs):
#         waiting_trucks = Trucks.objects.filter(is_waiting=True).values_list('truck_name', flat=True)
#         goods_on_road = GoodsOnRoad.objects.filter(is_on_truck=True).values()
#         grouped_goods = {}


#         for good in goods_on_road:

#             if good['truck_name'] not in grouped_goods:

#                 grouped_goods[good['truck_name']] = []


#             grouped_goods[good['truck_name']].append(good)

#         return JsonResponse(grouped_goods, safe=False)

# class ApproveWaitingTruckView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             data = json.loads(request.body)
#             truck_name = data.get("truck_name")

#             if not truck_name:
#                 return JsonResponse({'error': "Truck name cannot be empty."}, status=400)

#             truck = Trucks.objects.get(truck_name=truck_name, is_waiting=True)

#             if not truck:
#                 return JsonResponse({'error': "No truck found with the given name on waiting list"}, status=404)

#             truck.is_waiting = False
#             truck.is_ordered = True
#             truck.save()
            
#             goods_on_road = GoodsOnRoad.objects.filter(truck_name=truck_name, is_on_truck=True)

#             for good in goods_on_road:
#                 good.is_on_truck = False
#                 good.is_on_road = True
#                 good.save()

#             return JsonResponse({'message': f"Truck with name: {truck_name} is successfully ordered."}, status=200)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

# class TrucksOnRoadView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def get(self, request, *args, **kwargs):
#         trucks_on_road = Trucks.objects.filter(is_ordered=True).values_list('truck_name', flat=True)
#         goods_on_road = GoodsOnRoad.objects.filter(is_on_road = True).values()
#         grouped_goods = {}

#         for good in goods_on_road:
#             if good['truck_name'] not in grouped_goods:
#                 grouped_goods[good['truck_name']] = []

#             grouped_goods[good['truck_name']].append(good)

#         return JsonResponse(grouped_goods, safe=False)
    
# class EditGoodsOnRoadView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             data = json.loads(request.body)

#             product_code = data.get('product_code')
#             good_on_road = GoodsOnRoad.objects.get(product_code=product_code)

#             # Update decided_order
#             new_decided_order = float(data.get('new_decided_order'))
#             if new_decided_order is not None and new_decided_order >= 0:
#                 good_on_road.decided_order = new_decided_order
#             else:
#                 return JsonResponse({'error': "Decided Order cannot be negative or empty."}, status=400)

#             good_on_road.save()
#             return JsonResponse({'message': "Your changes have been successfully saved."}, status=200)

#         except GoodsOnRoad.DoesNotExist:
#             return JsonResponse({'error': "Good not found."}, status=400)

#         except ValueError as e:
#             return JsonResponse({'error': str(e)}, status=400)

#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

    

# class ApproveArrivedTruckView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         try:
#             data = json.loads(request.body)
#             truck_name = data.get("truck_name")

#             if not truck_name:
#                 return JsonResponse({'error': "Truck name cannot be empty."}, status=400)

#             truck = Trucks.objects.filter(truck_name=truck_name, is_ordered=True).first()
#             if not truck:
#                 return JsonResponse({'error': "No truck found with the given name and is_ordered=True."}, status=404)

#             goods_on_road = GoodsOnRoad.objects.filter(truck_name=truck_name, is_on_road=True)
#             for good in goods_on_road:
#                 try:
#                     warehouse_product = Warehouse.objects.get(product_code=good.product_code)
#                 except Warehouse.DoesNotExist:
#                    return JsonResponse({'error': f"Product with code {good.product_code} does not exist in the warehouse."}, status=500)

#             truck.is_ordered = False
#             truck.is_arrived = True
#             truck.save()

#             goods_on_road = GoodsOnRoad.objects.filter(truck_name=truck_name, is_on_road=True)
#             for good in goods_on_road:
#                 good.is_on_road = False
#                 good.is_arrived = True
#                 good.save()

#                 warehouse_product = Warehouse.objects.get(product_code=good.product_code)
#                 if warehouse_product:
#                     warehouse_product.stock += good.decided_order
#                     warehouse_product.save()

#             return JsonResponse({'message': "Truck, related GoodsOnRoad objects, and Warehouse stock updated successfully."}, status=200)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

# # endregion


# # region Notifications

# @receiver(pre_save, sender=OrderList)
# def create_notification_order_list(sender, instance, created, **kwargs):
#     notification_order_list = NotificationsOrderList(
#         current_date=instance.current_date,
#         product_code=instance.product_code,
#         is_active=instance.is_active,
#         order_avrg=instance.order_avrg,
#         order_exp=instance.order_exp,
#         order_holt=instance.order_holt,
#     )
#     notification_order_list.save()

# class NotificationsView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def get(self, request, *args, **kwargs):
#         notifications = NotificationsOrderList.objects.filter(is_active=True).values()
#         notification_data = [
#             [n['id'], n['current_date'].strftime('%Y-%m-%d'), n['product_code'], n['order_avrg'], n['order_exp'], n['order_holt']]
#             for n in notifications
#         ]
#         return JsonResponse(notification_data, safe=False)

# class DeleteNotificationView(APIView):
#     permission_classes = (IsAuthenticated,)
#     authentication_classes = (JWTAuthentication,)

#     def handle_exception(self, exc):
#         if isinstance(exc, (NotAuthenticated, PermissionDenied)):
#             return JsonResponse({'error': _("You do not have permission to perform this action.")}, status=400)

#         return super().handle_exception(exc)

#     def post(self, request, *args, **kwargs):
#         notification_id = request.data.get('id')
#         if not notification_id:
#             return JsonResponse({"error": "Missing 'id' parameter"}, status=400)

#         try:
#             notification = NotificationsOrderList.objects.get(id=notification_id)
#         except NotificationsOrderList.DoesNotExist:
#             return JsonResponse({"error": f"No notification found with id: {notification_id}"}, status=404)

#         notification.is_active = False
#         notification.save()

#         return JsonResponse({"success": f"Notification with id {notification_id} has been deactivated"})



# # endregion








